from fastapi import FastAPI, Request, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse, Response, JSONResponse
from anthropic import Anthropic
from dotenv import load_dotenv
from pydantic import BaseModel, Field
from typing import List
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import json, sqlite3, asyncio, io, smtplib, ssl, os, hmac, hashlib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import date
import stripe
import qrcode
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config import MANAGER_PASSWORD, ADMIN_PASSWORD, INVITE_CODE, PLAN_LIMITS, SECURE_COOKIES, DATABASE_PATH, CRON_SECRET, GUEST_TOKEN_SECRET
from database import (
    init_db, save_message, get_messages,
    create_hotel, get_hotel, get_all_hotels, get_hotel_messages,
    save_hotel_message, mark_hotel_read, mark_all_read,
    update_hotel, verify_password, migrate_password_if_needed,
    get_room_messages, get_new_messages, save_staff_message,
    save_rating, get_hotel_avg_rating, get_monthly_message_count,
    delete_hotel,
    save_guest, get_guests, update_guest_room, update_guest_status,
    get_active_guests_count, get_guest_by_room, mark_guest_reviewed,
    get_recent_ratings,
    save_telegram_msg_id, get_room_by_telegram_msg_id,
    get_daily_digest_data,
    detect_request_category, save_request, get_requests,
    update_request_status, delete_request, get_pending_requests_count,
    auto_checkout_overdue,
    get_room_notes, save_room_note, delete_room_note,
    create_staff, get_staff_list, get_staff_by_credentials,
    get_staff_by_id, update_staff_password, delete_staff,
    update_hotel_stripe, get_hotel_by_stripe_customer,
    create_owner, get_owner_by_email, get_owner_by_id, get_all_owners,
    assign_hotel_to_owner, remove_hotel_from_owner,
    get_owner_hotels, get_hotel_owner_ids,
    create_service, get_services, update_service, delete_service,
    save_staff_msg, get_staff_msgs, get_staff_msgs_since, STAFF_CHANNELS,
    generate_room_numbers,
)
from telegram import send_urgent_alert, send_telegram, set_webhook
from templates.chat_html import get_chat_html
from templates.checkin_html import get_checkin_html
from templates.dashboard_html import get_dashboard_html
from templates.admin_html import get_admin_html, get_admin_login_html
from templates.other_html import get_login_html, get_register_html, get_edit_html
from templates.landing_html import get_landing_html
from templates.owner_html import get_owner_login_html, get_owner_dashboard_html
from templates.public_html import get_public_page_html
from templates.widget_html import get_widget_html
from templates.buffet_html import get_buffet_html
from buffet import analyze_buffet_photo, save_buffet_scan, get_buffet_latest, get_buffet_history

load_dotenv()

# ===== STRIPE CONFIG =====
stripe.api_key = os.getenv("STRIPE_SECRET_KEY", "")
STRIPE_WEBHOOK_SECRET = os.getenv("STRIPE_WEBHOOK_SECRET", "")
# Map Stripe price IDs → internal plan names (must match PLAN_LIMITS keys)
STRIPE_PRICE_TO_PLAN: dict[str, str] = {
    p: name for p, name in [
        (os.getenv("STRIPE_PRICE_BASIC", ""),      "starter"),
        (os.getenv("STRIPE_PRICE_PRO", ""),         "pro"),
        (os.getenv("STRIPE_PRICE_ENTERPRISE", ""),  "premium"),
    ] if p
}

# ===== RATE LIMITER =====
limiter = Limiter(key_func=get_remote_address)
app = FastAPI()
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

client = Anthropic()


# ===== GUEST TOKEN HELPERS =====
# Guest cookie: name=guest_{slug}, value="{room}:{hmac}"
# The HMAC signs slug+room so a guest in room 101 can't forge a token for 102.

def _make_guest_token(slug: str, room: str) -> str:
    key = GUEST_TOKEN_SECRET.encode()
    msg = f"{slug}:{room}".encode()
    return hmac.new(key, msg, hashlib.sha256).hexdigest()

def _is_guest_for_room(request: Request, slug: str, room: str) -> bool:
    cookie_val = request.cookies.get(f"guest_{slug}", "")
    if not cookie_val or ":" not in cookie_val:
        return False
    stored_room, stored_hmac = cookie_val.split(":", 1)
    expected = _make_guest_token(slug, room)
    return stored_room == room and hmac.compare_digest(stored_hmac, expected)


# ===== EMAIL HELPER =====
async def send_email(hotel: dict, subject: str, body_html: str) -> bool:
    """Send an HTML email using hotel's own SMTP config. Returns True on success."""
    smtp_host = hotel.get("smtp_host", "").strip()
    notify_email = hotel.get("notify_email", "").strip()
    if not smtp_host or not notify_email:
        return False  # SMTP not configured — silently skip
    smtp_port = int(hotel.get("smtp_port") or 587)
    smtp_user = hotel.get("smtp_user", "").strip()
    smtp_pass = hotel.get("smtp_pass", "")
    from_addr = hotel.get("smtp_from", "").strip() or smtp_user

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = f"{hotel['name']} <{from_addr}>"
    msg["To"] = notify_email
    msg.attach(MIMEText(body_html, "html", "utf-8"))

    def _send():
        try:
            if smtp_port == 465:
                ctx = ssl.create_default_context()
                with smtplib.SMTP_SSL(smtp_host, smtp_port, context=ctx, timeout=10) as s:
                    if smtp_user:
                        s.login(smtp_user, smtp_pass)
                    s.sendmail(from_addr, notify_email, msg.as_string())
            else:
                with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as s:
                    s.ehlo()
                    s.starttls()
                    s.ehlo()
                    if smtp_user:
                        s.login(smtp_user, smtp_pass)
                    s.sendmail(from_addr, notify_email, msg.as_string())
            return True
        except Exception as e:
            print(f"[email] send failed for {hotel['slug']}: {e}")
            return False

    return await asyncio.to_thread(_send)


HOTEL_INFO = """
Ты AI консьерж отеля SmartStay Resort 5* в Анталии, Турция.
Отвечай ТОЛЬКО на языке на котором пишет гость.
Отвечай коротко, дружелюбно и по делу.
Если не знаешь ответ — скажи "Уточню у персонала".
"""

init_db()

# ===== PYDANTIC MODELS =====
class ChatRequest(BaseModel):
    message: str = Field(..., max_length=5000)
    history: List[dict] = []
    room: str = Field(default="101", max_length=20)

class RegisterRequest(BaseModel):
    slug: str
    name: str
    password: str
    info: str
    tg_token: str = ""
    tg_chat: str = ""
    invite_code: str = ""
    room_count: int = 30
    room_start: int = 101
    rooms_per_floor: int = 0

class LoginRequest(BaseModel):
    password: str

class HotelLoginRequest(BaseModel):
    slug: str
    password: str

class UpdateRequest(BaseModel):
    name: str
    info: str
    password: str = ""
    tg_token: str = ""
    tg_chat: str = ""
    booking_url: str = ""
    ai_name: str = ""
    smtp_host: str = ""
    smtp_port: int = 587
    smtp_user: str = ""
    smtp_pass: str = ""
    smtp_from: str = ""
    notify_email: str = ""
    default_language: str = "auto"
    supported_languages: str = "en,ru,tr,ar,de,fr"
    photo_url: str = ""
    page_description: str = ""
    amenities: str = ""
    rooms_per_floor: int = 0

class RatingRequest(BaseModel):
    room: str
    rating: int

class CheckinRequest(BaseModel):
    first_name: str
    last_name: str
    passport: str
    nationality: str = ""
    room: str
    check_in: str
    check_out: str

class GuestStatusRequest(BaseModel):
    status: str
    room: str = ""
    notes: str = ""

class ReplyRequest(BaseModel):
    message: str

class AdminLoginRequest(BaseModel):
    password: str

# ===== HEALTH CHECK =====
@app.get("/favicon.ico")
def favicon():
    # Simple gold hotel emoji favicon as SVG
    svg = b'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32"><rect width="32" height="32" rx="6" fill="#C9A84C"/><text x="16" y="23" font-size="18" text-anchor="middle">&#x1F3E8;</text></svg>'
    return Response(content=svg, media_type="image/svg+xml", headers={"Cache-Control": "public, max-age=86400"})

@app.get("/health")
def health():
    """Railway health check — also verifies DB is reachable."""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.execute("SELECT 1")
        conn.close()
        return {"status": "ok", "db": "ok"}
    except Exception as e:
        return JSONResponse({"status": "error", "detail": str(e)}, status_code=500)

# ===== ГЛАВНАЯ =====
@app.get("/", response_class=HTMLResponse)
def home():
    return get_landing_html()

# ===== РЕГИСТРАЦИЯ =====
@app.get("/register", response_class=HTMLResponse)
def register_page():
    return get_register_html()

@app.get("/api/register-config")
def register_config():
    """Tells the frontend whether an invite code is required."""
    return {"invite_required": bool(INVITE_CODE)}

@app.post("/api/register")
def api_register(data: RegisterRequest):
    # Invite code gate — empty INVITE_CODE means open registration (dev/pilot mode)
    if INVITE_CODE and data.invite_code.strip() != INVITE_CODE:
        return {"ok": False, "error": "Geçersiz davet kodu"}

    slug = data.slug.strip()
    name = data.name.strip()
    password = data.password.strip()
    info = data.info.strip()

    if not all([slug, name, password, info]):
        return {"ok": False, "error": "Zorunlu alanlar eksik"}
    if get_hotel(slug):
        return {"ok": False, "error": "Bu slug zaten kullanılıyor"}

    room_count = max(1, min(data.room_count, 500))  # clamp 1..500
    room_start = max(1, min(data.room_start, 9000))
    rooms_per_floor = max(0, min(data.rooms_per_floor, 100))

    create_hotel(slug, name, password, info,
                 data.tg_token.strip(), data.tg_chat.strip(),
                 plan="trial", room_count=room_count, room_start=room_start,
                 rooms_per_floor=rooms_per_floor)
    return {"ok": True, "slug": slug}

# ===== LOGIN =====
@app.get("/login", response_class=HTMLResponse)
def login_page():
    return get_login_html()

@app.post("/api/login")
@limiter.limit("5/minute")
def api_login(data: LoginRequest, request: Request, response: Response):
    if data.password == MANAGER_PASSWORD:
        response.set_cookie("manager_auth", "yes", max_age=86400,
                            httponly=True, secure=SECURE_COOKIES, samesite="lax")
        return {"ok": True}
    return {"ok": False}

@app.get("/hotel/{slug}/login", response_class=HTMLResponse)
def hotel_login_page(slug: str):
    hotel = get_hotel(slug)
    if not hotel:
        return HTMLResponse("❌ Otel bulunamadı", status_code=404)
    return get_login_html(hotel["name"])

@app.post("/api/hotel-login")
@limiter.limit("5/minute")
def api_hotel_login(data: HotelLoginRequest, request: Request):
    hotel = get_hotel(data.slug)
    if hotel and verify_password(data.password, hotel["password"]):
        migrate_password_if_needed(data.slug, data.password, hotel["password"])
        response = JSONResponse({"ok": True})
        response.set_cookie(
            key=f"auth_{data.slug}",
            value="yes",
            max_age=86400,
            httponly=True,
            secure=SECURE_COOKIES,
            samesite="lax",
            path="/"
        )
        return response
    return JSONResponse({"ok": False})

# ===== LOGOUT =====
@app.get("/hotel/{slug}/logout")
def hotel_logout(slug: str, response: Response):
    response.delete_cookie(f"auth_{slug}", path="/")
    return RedirectResponse(f"/hotel/{slug}/login", status_code=302)

# ===== DASHBOARD =====
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    if request.cookies.get("manager_auth") != "yes":
        return RedirectResponse("/login")
    return get_dashboard_html()

@app.get("/hotel/{slug}/dashboard", response_class=HTMLResponse)
def hotel_dashboard(slug: str, request: Request):
    hotel = get_hotel(slug)
    if not hotel:
        return HTMLResponse("❌ Otel bulunamadı", status_code=404)
    ok, role = get_auth_role(request, slug)
    if not ok:
        return RedirectResponse(f"/hotel/{slug}/login", status_code=302)
    return get_dashboard_html(hotel["name"])

# ===== HOTEL CHAT =====
@app.get("/hotel/{slug}", response_class=HTMLResponse)
def hotel_chat(slug: str, room: str = "101"):
    hotel = get_hotel(slug)
    if not hotel:
        return HTMLResponse("❌ Otel bulunamadı", status_code=404)
    # Sanitise room before signing — reject anything non-alphanumeric
    room = room.strip()
    if not room.isalnum() or len(room) > 20:
        room = "101"
    html = get_chat_html(hotel["name"], slug, hotel.get("default_language") or "en")
    resp = HTMLResponse(html)
    resp.set_cookie(
        key=f"guest_{slug}",
        value=f"{room}:{_make_guest_token(slug, room)}",
        httponly=True,
        secure=SECURE_COOKIES,
        samesite="lax",  # lax allows first-visit from QR scan (top-level navigation)
        max_age=86400 * 7,
    )
    return resp

# ===== PWA =====
@app.get("/hotel/{slug}/manifest.json")
def hotel_manifest(slug: str):
    hotel = get_hotel(slug)
    name = hotel["name"] if hotel else "SmartStay"
    short = name[:12]
    return JSONResponse({
        "name": name,
        "short_name": short,
        "description": "Hotel AI Concierge",
        "start_url": f"/hotel/{slug}",
        "scope": f"/hotel/{slug}",
        "display": "standalone",
        "background_color": "#141414",
        "theme_color": "#C9A84C",
        "orientation": "portrait",
        "icons": [
            {"src": f"/hotel/{slug}/icon.svg", "sizes": "any", "type": "image/svg+xml", "purpose": "any"}
        ]
    })

@app.get("/hotel/{slug}/icon.svg")
def hotel_icon(slug: str):
    hotel = get_hotel(slug)
    name = hotel["name"] if hotel else "SmartStay"
    initials = "".join(w[0].upper() for w in name.split()[:2]) or "S"
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 512 512">
  <rect width="512" height="512" rx="80" fill="#141414"/>
  <rect x="16" y="16" width="480" height="480" rx="72" fill="none" stroke="#C9A84C" stroke-width="8"/>
  <text x="256" y="300" font-size="220" font-family="system-ui,sans-serif" font-weight="700"
        text-anchor="middle" dominant-baseline="middle" fill="#C9A84C">{initials}</text>
</svg>"""
    return Response(svg, media_type="image/svg+xml",
                    headers={"Cache-Control": "public, max-age=86400"})

@app.get("/hotel/{slug}/sw.js")
def hotel_sw(slug: str):
    js = f"""const CACHE = 'ss-{slug}-v1';
const SHELL = ['/hotel/{slug}'];

self.addEventListener('install', e => {{
  e.waitUntil(caches.open(CACHE).then(c => c.addAll(SHELL)));
  self.skipWaiting();
}});
self.addEventListener('activate', e => {{
  e.waitUntil(caches.keys().then(ks =>
    Promise.all(ks.filter(k => k !== CACHE).map(k => caches.delete(k)))));
  self.clients.claim();
}});
self.addEventListener('fetch', e => {{
  if (e.request.method !== 'GET') return;
  if (e.request.url.includes('/api/')) return;
  e.respondWith(caches.match(e.request).then(r => r || fetch(e.request)));
}});"""
    return Response(js, media_type="application/javascript",
                    headers={"Cache-Control": "no-cache"})

# ===== PUBLIC HOTEL PAGE =====
@app.get("/hotel/{slug}/page", response_class=HTMLResponse)
def hotel_public_page(slug: str):
    hotel = get_hotel(slug)
    if not hotel:
        return HTMLResponse("❌ Otel bulunamadı", status_code=404)
    avg_rating, rating_count = get_hotel_avg_rating(slug)
    # Fetch recent ratings (last 10)
    conn = sqlite3.connect(DATABASE_PATH)
    recent_rows = conn.execute(
        "SELECT rating, created_at FROM ratings WHERE hotel_slug=? ORDER BY id DESC LIMIT 10",
        (slug,)
    ).fetchall()
    conn.close()
    recent_reviews = [{"rating": r[0], "created_at": r[1]} for r in recent_rows]
    return get_public_page_html(hotel, avg_rating, rating_count, recent_reviews)


@app.get("/hotel/{slug}/widget", response_class=HTMLResponse)
def hotel_widget(slug: str):
    """Minimal chat UI for iframe embedding."""
    hotel = get_hotel(slug)
    if not hotel:
        return HTMLResponse("❌ Otel bulunamadı", status_code=404)
    return get_widget_html(hotel["name"], slug, hotel.get("ai_name") or "AI Asistan")


@app.get("/hotel/{slug}/embed.js")
def hotel_embed_js(slug: str, request: Request):
    """Self-contained embed script. Drop <script src='.../embed.js'></script> on any page."""
    hotel = get_hotel(slug)
    label = hotel["name"] if hotel else "SmartStay"
    base = str(request.base_url).rstrip("/")
    widget_url = f"{base}/hotel/{slug}/widget"
    js = f"""(function(){{
  if(document.getElementById('ss-embed-{slug}'))return;
  var BASE='{base}',SLUG='{slug}',LABEL={repr(label)};

  var style=document.createElement('style');
  style.textContent=`
    #ss-bubble-{slug}{{
      position:fixed;bottom:24px;right:24px;z-index:99999;
      width:60px;height:60px;border-radius:50%;
      background:linear-gradient(135deg,#C9A84C,#E8C96A);
      box-shadow:0 4px 20px rgba(201,168,76,0.4);
      display:flex;align-items:center;justify-content:center;
      font-size:28px;cursor:pointer;
      transition:transform 0.2s,box-shadow 0.2s;
      border:none;outline:none;
    }}
    #ss-bubble-{slug}:hover{{transform:scale(1.08);box-shadow:0 6px 28px rgba(201,168,76,0.5);}}
    #ss-frame-wrap-{slug}{{
      position:fixed;bottom:96px;right:24px;z-index:99998;
      width:380px;height:600px;border-radius:20px;overflow:hidden;
      box-shadow:0 20px 60px rgba(0,0,0,0.4);
      transition:opacity 0.2s,transform 0.2s;
      opacity:0;pointer-events:none;transform:translateY(12px);
    }}
    #ss-frame-wrap-{slug}.open{{opacity:1;pointer-events:all;transform:translateY(0);}}
    #ss-frame-{slug}{{width:100%;height:100%;border:none;border-radius:20px;}}
    @media(max-width:480px){{
      #ss-frame-wrap-{slug}{{
        bottom:0;right:0;width:100%;height:80vh;border-radius:20px 20px 0 0;
      }}
    }}
  `;
  document.head.appendChild(style);

  var bubble=document.createElement('button');
  bubble.id='ss-bubble-{slug}';
  bubble.innerHTML='🛎️';
  bubble.title='Чат с консьержем';

  var wrap=document.createElement('div');
  wrap.id='ss-frame-wrap-{slug}';
  wrap.innerHTML='<iframe id="ss-frame-{slug}" src="{widget_url}" allow="microphone" loading="lazy"></iframe>';

  var open=false;
  bubble.onclick=function(){{
    open=!open;
    wrap.classList.toggle('open',open);
    bubble.innerHTML=open?'✕':'🛎️';
  }};

  document.body.appendChild(wrap);
  document.body.appendChild(bubble);
  document.getElementById('ss-embed-{slug}')|| (function(){{
    var m=document.createElement('meta');m.id='ss-embed-{slug}';
    document.head.appendChild(m);
  }})();
}})();"""
    return Response(js, media_type="application/javascript",
                    headers={"Cache-Control": "public, max-age=3600"})


@app.post("/hotel/{slug}/chat")
@limiter.limit("30/minute")
async def hotel_chat_api(request: Request, slug: str, data: ChatRequest):
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Hotel not found"}, status_code=404)

    message = data.message.strip()
    room = data.room.strip() or "101"
    history = data.history

    if not message:
        return JSONResponse({"error": "Empty message"}, status_code=400)
    if not room.isalnum():
        return JSONResponse({"error": "Geçersiz oda numarası"}, status_code=400)

    # --- Plan limit check ---
    plan = hotel.get("plan", "trial")
    limit = PLAN_LIMITS.get(plan, PLAN_LIMITS["trial"])
    if limit != -1:
        used = get_monthly_message_count(slug)
        if used >= limit:
            return JSONResponse({
                "error": f"Bu ay mesaj limitine ulaşıldı ({limit} mesaj). "
                         f"Lütfen otel yöneticisiyle iletişime geçin."
            }, status_code=429)

    # Load guest context once — used for system prompt, request tracker, etc.
    guest = get_guest_by_room(slug, room)

    priority = save_hotel_message(slug, room, "user", message)

    # Auto-detect service requests and save to tracker
    req_category = detect_request_category(message)
    if req_category:
        guest_name = f"{guest['first_name']} {guest['last_name']}" if guest else ""
        save_request(slug, room, guest_name, req_category, message)

    if priority == "urgent":
        async def _send_urgent_and_map():
            msg_id = await send_urgent_alert(
                hotel_name=hotel["name"], room=room, message=message,
                token=hotel.get("telegram_token"),
                chat_id=hotel.get("telegram_chat_id")
            )
            if msg_id:
                save_telegram_msg_id(slug, room, msg_id)
        asyncio.create_task(_send_urgent_and_map())
    elif hotel.get("telegram_token") and hotel.get("telegram_chat_id"):
        # Non-urgent messages: also send to Telegram and save mapping
        async def _send_normal_and_map():
            from datetime import datetime as _dt
            tg_text = (
                f"💬 <b>{hotel['name']}</b>\n"
                f"🚪 Oda {room} · {_dt.now().strftime('%H:%M')}\n"
                f"{message}\n\n"
                f"<i>↩️ Yanıt vermek için bu mesaja cevap verin.</i>"
            )
            msg_id = await send_telegram(
                tg_text,
                token=hotel.get("telegram_token"),
                chat_id=hotel.get("telegram_chat_id")
            )
            if msg_id:
                save_telegram_msg_id(slug, room, msg_id)
        asyncio.create_task(_send_normal_and_map())

    # Email notification (independent of Telegram)
    if hotel.get("notify_email") and hotel.get("smtp_host"):
        from datetime import datetime as _dt2
        _ts = _dt2.now().strftime("%H:%M")
        _guest_label = ""
        if guest:
            _guest_label = f"<br><b>Misafir:</b> {guest['first_name']} {guest['last_name']}"
        if priority == "urgent":
            _email_subject = f"🔴 ACİL — {hotel['name']} Oda {room} [{_ts}]"
            _email_body = f"""
<div style="font-family:sans-serif;max-width:600px">
  <div style="background:#c0392b;color:#fff;padding:16px 20px;border-radius:8px 8px 0 0">
    <h2 style="margin:0">🔴 ACİL Misafir Mesajı</h2>
  </div>
  <div style="border:1px solid #e0e0e0;border-top:none;padding:20px;border-radius:0 0 8px 8px">
    <p><b>Otel:</b> {hotel['name']}<br>
       <b>Oda:</b> {room}{_guest_label}<br>
       <b>Saat:</b> {_ts}</p>
    <div style="background:#fff3f3;border-left:4px solid #c0392b;padding:12px 16px;border-radius:4px;margin:12px 0">
      <b>Mesaj:</b><br>{message}
    </div>
    <p style="color:#888;font-size:12px">SmartStay AI tarafından gönderildi</p>
  </div>
</div>"""
        else:
            _email_subject = f"💬 Yeni Mesaj — {hotel['name']} Oda {room} [{_ts}]"
            _email_body = f"""
<div style="font-family:sans-serif;max-width:600px">
  <div style="background:#C9A84C;color:#fff;padding:16px 20px;border-radius:8px 8px 0 0">
    <h2 style="margin:0">💬 Yeni Misafir Mesajı</h2>
  </div>
  <div style="border:1px solid #e0e0e0;border-top:none;padding:20px;border-radius:0 0 8px 8px">
    <p><b>Otel:</b> {hotel['name']}<br>
       <b>Oda:</b> {room}{_guest_label}<br>
       <b>Saat:</b> {_ts}</p>
    <div style="background:#fafaf5;border-left:4px solid #C9A84C;padding:12px 16px;border-radius:4px;margin:12px 0">
      <b>Mesaj:</b><br>{message}
    </div>
    <p style="color:#888;font-size:12px">SmartStay AI tarafından gönderildi</p>
  </div>
</div>"""
        asyncio.create_task(send_email(hotel, _email_subject, _email_body))

    history.append({"role": "user", "content": message})

    # ── Build rich system prompt with guest context ─────────────────────────
    from datetime import datetime as _dt, date as _date
    today = _date.today()

    guest_lines = []
    if guest:
        full_name = f"{guest['first_name']} {guest['last_name']}"
        guest_lines.append(f"Misafir adı: {full_name}")
        try:
            ci = _dt.strptime(guest["check_in"], "%Y-%m-%d").date()
            co = _dt.strptime(guest["check_out"], "%Y-%m-%d").date()
            days_in = (today - ci).days or 1
            days_left = (co - today).days
            guest_lines.append(f"Check-in: {guest['check_in']}  Check-out: {guest['check_out']}")
            guest_lines.append(f"Konaklama: {days_in}. gün — çıkışa {days_left} gün kaldı")
        except ValueError:
            pass

    # Pending requests for this room
    room_requests = [r for r in get_requests(slug) if r["room"] == room and r["status"] != "resolved"]
    if room_requests:
        req_summary = ", ".join(
            f"{r['category'].replace('_',' ')} ({r['status']})" for r in room_requests[:3]
        )
        guest_lines.append(f"Açık talepler: {req_summary}")

    guest_block = ("\n\nMisafir bilgileri:\n" + "\n".join(guest_lines)) if guest_lines else ""

    ai_name = hotel.get("ai_name") or "AI Asistan"
    default_lang = hotel.get("default_language") or "auto"
    supported_langs = hotel.get("supported_languages") or "en,ru,tr,ar,de,fr"

    if default_lang == "auto":
        lang_instruction = (
            "LANGUAGE RULE (critical): Detect the language of the guest's last message "
            "and always reply in that exact same language. "
            "Never mix languages in one reply. "
            "You support all world languages — respond fluently in whatever language the guest uses: "
            "Russian, English, Turkish, Uzbek, Arabic, Chinese, German, French, Spanish, Italian, "
            "Portuguese, Japanese, Korean, Hindi, Persian, Azerbaijani, Kazakh, Kyrgyz, Turkmen, "
            "Ukrainian, Polish, Dutch, Indonesian, Malay, Romanian, Czech, Hungarian, and any other language. "
            "Always match the guest's language exactly."
        )
    else:
        lang_map = {
            "en": "English", "ru": "Russian", "tr": "Turkish", "uz": "Uzbek",
            "ar": "Arabic", "zh": "Chinese", "de": "German", "fr": "French",
            "es": "Spanish", "it": "Italian", "pt": "Portuguese",
            "ja": "Japanese", "ko": "Korean", "hi": "Hindi", "fa": "Persian",
            "az": "Azerbaijani", "kk": "Kazakh", "ky": "Kyrgyz", "tk": "Turkmen",
            "uk": "Ukrainian", "pl": "Polish", "nl": "Dutch", "id": "Indonesian",
            "ms": "Malay", "ro": "Romanian", "cs": "Czech", "hu": "Hungarian",
        }
        lang_name = lang_map.get(default_lang, default_lang.upper())
        lang_instruction = (
            f"LANGUAGE RULE: Always respond in {lang_name}, regardless of what language the guest writes in. "
            f"This hotel has set {lang_name} as the required response language."
        )

    system = (
        f"You are '{ai_name}', the AI concierge assistant for {hotel['name']} hotel.\n\n" +
        hotel["info"] +
        f"\n\nHotel: {hotel['name']}\nRoom: {room}. Do not ask for the room number again." +
        guest_block +
        f"\n\nIntroduce yourself as '{ai_name}'. Be warm, personalised and professional. "
        "Accept all requests and confirm they will be forwarded to staff. "
        f"\n\n{lang_instruction}"
    )

    def generate():
        try:
            full_response = []
            with client.messages.stream(
                model="claude-sonnet-4-5",
                max_tokens=500,
                system=system,
                messages=history
            ) as stream:
                for text in stream.text_stream:
                    full_response.append(text)
                    yield f"data: {json.dumps({'text': text})}\n\n"
            bot_msg = "".join(full_response)
            save_hotel_message(slug, room, "bot", bot_msg)
            yield f"data: {json.dumps({'done': True})}\n\n"
        except Exception as e:
            print(f"[Claude stream error] {slug}/{room}: {e}")
            yield f"data: {json.dumps({'error': 'Üzgünüm, şu an bağlanamıyorum. Lütfen bir dakika sonra tekrar deneyin.'})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")

# ===== MAIN CHAT =====
@app.post("/chat")
@limiter.limit("30/minute")
async def chat(request: Request, data: ChatRequest):
    message = data.message.strip()
    room = data.room.strip() or "101"
    history = data.history

    if not message:
        return JSONResponse({"error": "Empty message"}, status_code=400)
    if not room.isalnum():
        return JSONResponse({"error": "Geçersiz oda numarası"}, status_code=400)

    save_message(room, "user", message)
    history.append({"role": "user", "content": message})

    def generate():
        try:
            full_response = []
            with client.messages.stream(
                model="claude-sonnet-4-5",
                max_tokens=500,
                system=HOTEL_INFO + f"\n\nНомер гостя: {room}.",
                messages=history
            ) as stream:
                for text in stream.text_stream:
                    full_response.append(text)
                    yield f"data: {json.dumps({'text': text})}\n\n"
            save_message(room, "bot", "".join(full_response))
            yield f"data: {json.dumps({'done': True})}\n\n"
        except Exception as e:
            print(f"[Claude stream error] demo/{room}: {e}")
            yield f"data: {json.dumps({'error': 'Üzgünüm, şu an bağlanamıyorum. Lütfen bir dakika sonra tekrar deneyin.'})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")

# ===== TRANSLATE =====
@app.post("/api/translate")
@limiter.limit("60/minute")
async def translate_text(request: Request):
    data = await request.json()
    text = (data.get("text") or "").strip()
    target = data.get("target") or "en"
    if not text:
        return JSONResponse({"error": "No text"}, status_code=400)
    lang_map = {
        "en": "English", "ru": "Russian", "tr": "Turkish", "uz": "Uzbek",
        "ar": "Arabic", "zh": "Chinese", "de": "German", "fr": "French",
        "es": "Spanish", "it": "Italian", "pt": "Portuguese",
        "ja": "Japanese", "ko": "Korean", "hi": "Hindi", "fa": "Persian",
        "az": "Azerbaijani", "kk": "Kazakh", "ky": "Kyrgyz", "tk": "Turkmen",
        "uk": "Ukrainian", "pl": "Polish", "nl": "Dutch", "id": "Indonesian",
    }
    lang_name = lang_map.get(target, "English")
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=500,
            messages=[{"role": "user", "content":
                f"Translate to {lang_name}. Return ONLY the translation, no explanations:\n\n{text}"}]
        )
        return {"translation": resp.content[0].text.strip()}
    except Exception as e:
        print(f"[translate error] {e}")
        return JSONResponse({"error": "Bir hata oluştu, lütfen tekrar deneyin"}, status_code=500)

# ===== SELF CHECK-IN =====
@app.get("/hotel/{slug}/checkin", response_class=HTMLResponse)
def checkin_page(slug: str):
    hotel = get_hotel(slug)
    if not hotel:
        return HTMLResponse("❌ Otel bulunamadı", status_code=404)
    return get_checkin_html(hotel["name"], hotel.get("default_language") or "en")

@app.post("/api/hotel/{slug}/checkin")
@limiter.limit("10/minute")
async def api_checkin(request: Request, slug: str, data: CheckinRequest):
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"ok": False, "error": "Hotel not found"}, status_code=404)

    # Basic validation
    if not all([data.first_name.strip(), data.last_name.strip(), data.passport.strip()]):
        return JSONResponse({"ok": False, "error": "Ad, soyad ve pasaport zorunludur"})
    if not data.room.strip():
        return JSONResponse({"ok": False, "error": "Oda numarası zorunludur"})
    if data.check_out <= data.check_in:
        return JSONResponse({"ok": False, "error": "Çıkış tarihi giriş tarihinden sonra olmalı"})

    # Claude validation — quick sanity check
    validation_ok = True
    validation_error = ""
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=60,
            messages=[{
                "role": "user",
                "content": (
                    "Validate this hotel check-in. Reply ONLY 'VALID' or 'INVALID: [short reason in English]'.\n"
                    f"Name: {data.first_name} {data.last_name}\n"
                    f"Passport: {data.passport}\n"
                    f"Check-in: {data.check_in}, Check-out: {data.check_out}\n"
                    f"Room: {data.room}"
                )
            }]
        )
        result = resp.content[0].text.strip()
        if result.startswith("INVALID"):
            validation_ok = False
            validation_error = result.replace("INVALID:", "").strip()
    except Exception as e:
        print(f"[Checkin validation error] {e}")
        # Fail open — don't block guest if Claude is down

    if not validation_ok:
        return JSONResponse({"ok": False, "error": f"Bilgiler geçersiz: {validation_error}"})

    # Save to DB
    guest_id = save_guest(
        hotel_slug=slug,
        room=data.room.strip(),
        first_name=data.first_name.strip(),
        last_name=data.last_name.strip(),
        passport=data.passport.strip(),
        nationality=data.nationality.strip(),
        check_in=data.check_in,
        check_out=data.check_out,
    )

    # Set status to checked_in immediately
    update_guest_status(guest_id, "checked_in")

    # Telegram notification to manager
    tg_text = (
        f"🏨 <b>Yeni Check-in!</b>\n"
        f"👤 {data.first_name} {data.last_name} ({data.nationality})\n"
        f"🚪 Oda: {data.room}\n"
        f"📅 {data.check_in} → {data.check_out}\n"
        f"🪪 Pasaport: {data.passport}"
    )
    asyncio.create_task(
        send_telegram(
            tg_text,
            token=hotel.get("telegram_token"),
            chat_id=hotel.get("telegram_chat_id")
        )
    )

    return {"ok": True, "guest_id": guest_id, "room": data.room.strip()}

@app.get("/api/hotel/{slug}/guests")
def api_guests(slug: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    return get_guests(slug)

@app.post("/api/hotel/{slug}/guests/{guest_id}/status")
def api_guest_status(slug: str, guest_id: int, data: GuestStatusRequest, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    if data.status not in ("pending", "checked_in", "checked_out"):
        return JSONResponse({"error": "Invalid status"}, status_code=400)
    if data.room:
        update_guest_room(guest_id, data.room.strip())
    update_guest_status(guest_id, data.status, data.notes)
    return {"ok": True}

# ===== RATING =====
@app.get("/api/hotel/{slug}/room/{room}/review-needed")
def review_needed(slug: str, room: str, request: Request):
    """
    Returns whether a proactive review request should be shown to the guest.
    Accessible to: hotel manager/staff OR the guest who owns this room (via QR cookie).
    """
    if not get_auth_role(request, slug)[0] and not _is_guest_for_room(request, slug, room):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    guest = get_guest_by_room(slug, room)
    if not guest or guest["reviewed"]:
        return {"show": False}
    try:
        check_in_date = date.fromisoformat(guest["check_in"])
        days_stayed = (date.today() - check_in_date).days
        if days_stayed >= 1:
            return {"show": True, "guest_id": guest["id"],
                    "days": days_stayed, "name": guest["first_name"]}
    except Exception:
        pass
    return {"show": False}

@app.post("/hotel/{slug}/rate")
@limiter.limit("5/minute")
async def hotel_rate(request: Request, slug: str, data: RatingRequest):
    hotel = get_hotel(slug)
    if not hotel:
        return {"ok": False}
    if not 1 <= data.rating <= 5:
        return {"ok": False, "error": "Rating must be 1-5"}

    save_rating(slug, data.room, data.rating)

    # Mark guest as reviewed so we don't ask again
    guest = get_guest_by_room(slug, data.room)
    if guest:
        mark_guest_reviewed(guest["id"])

    booking_url = hotel.get("booking_url", "")

    if data.rating <= 3:
        # Low rating — alert manager immediately
        tg_text = (
            f"⚠️ <b>Düşük Puan Alındı!</b>\n"
            f"🏨 Otel: {hotel['name']}\n"
            f"🚪 Oda: {data.room}\n"
            f"⭐ Puan: {'★' * data.rating}{'☆' * (5 - data.rating)} ({data.rating}/5)\n"
            f"⏰ Şimdi harekete geçin — misafir ayrılmadan sorunu çözün!"
        )
        asyncio.create_task(send_telegram(
            tg_text,
            token=hotel.get("telegram_token"),
            chat_id=hotel.get("telegram_chat_id")
        ))
        return {"ok": True, "low": True}

    # High rating (4-5) — return Booking link if configured
    return {"ok": True, "low": False, "booking_url": booking_url}

# ===== STAFF REPLY =====
@app.post("/api/hotel/{slug}/room/{room}/reply")
async def staff_reply(slug: str, room: str, data: ReplyRequest, request: Request):
    if not get_auth_role(request, slug)[0]:
        return {"ok": False, "error": "Unauthorized"}
    message = data.message.strip()
    if not message:
        return {"ok": False, "error": "Empty message"}
    save_staff_message(slug, room, message)

    # Optionally forward reply to Telegram so guest chat & Telegram stay in sync
    hotel = get_hotel(slug)
    if hotel and hotel.get("telegram_token") and hotel.get("telegram_chat_id"):
        async def _fwd():
            from datetime import datetime as _dt
            tg_text = (
                f"📤 <b>Personel yanıtı</b> — {hotel['name']}\n"
                f"🚪 Oda {room} · {_dt.now().strftime('%H:%M')}\n"
                f"{message}"
            )
            await send_telegram(tg_text, token=hotel["telegram_token"],
                                chat_id=hotel["telegram_chat_id"])
        asyncio.create_task(_fwd())

    return {"ok": True}

# ===== ROOM CONVERSATION =====
@app.get("/api/hotel/{slug}/room/{room}/messages")
def room_messages(slug: str, room: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    return get_room_messages(slug, room)

@app.get("/api/hotel/{slug}/room/{room}/new-messages")
def room_new_messages(slug: str, room: str, since_id: int = 0, request: Request = None):
    if request and not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    return get_new_messages(slug, room, since_id)

# ===== PUBLIC GUEST CHAT HISTORY =====
@app.get("/api/hotel/{slug}/room/{room}/history")
def room_history_public(slug: str, room: str, request: Request):
    """
    Returns the last 50 messages for this room.
    Accessible to: hotel manager/staff OR the guest who owns this room (via QR cookie).
    """
    if not get_auth_role(request, slug)[0] and not _is_guest_for_room(request, slug, room):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Hotel not found"}, status_code=404)
    msgs = get_room_messages(slug, room, limit=50)
    return [
        {"id": m["id"], "role": m["role"], "message": m["message"],
         "created_at": m["created_at"]}
        for m in msgs
    ]


# ===== AUTO WELCOME MESSAGE =====
@app.post("/api/hotel/{slug}/room/{room}/welcome")
async def auto_welcome(slug: str, room: str):
    """
    Called on first chat open when history is empty and guest is checked in.
    Generates a personalised welcome via Claude Haiku and saves it as a bot message.
    Returns {"message": "..."} or {"skip": true} if no welcome needed.
    """
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Hotel not found"}, status_code=404)

    # Only generate if room has no messages yet
    existing = get_room_messages(slug, room, limit=1)
    if existing:
        return {"skip": True}

    guest = get_guest_by_room(slug, room)
    if not guest:
        return {"skip": True}

    from datetime import datetime as _dt, date as _date
    today = _date.today()
    first_name = guest["first_name"]
    try:
        co = _dt.strptime(guest["check_out"], "%Y-%m-%d").date()
        days_left = (co - today).days
        stay_info = f"check-out tarihi {guest['check_out']} ({days_left} gün kaldı)"
    except ValueError:
        stay_info = f"check-out: {guest['check_out']}"

    ai_name = hotel.get("ai_name") or "AI Asistan"
    default_lang = hotel.get("default_language") or "auto"
    lang_map = {
        "en": "English", "ru": "Russian", "tr": "Turkish", "uz": "Uzbek",
        "ar": "Arabic", "zh": "Chinese", "de": "German", "fr": "French",
        "es": "Spanish", "it": "Italian", "pt": "Portuguese",
        "ja": "Japanese", "ko": "Korean", "hi": "Hindi", "fa": "Persian",
        "az": "Azerbaijani", "kk": "Kazakh", "ky": "Kyrgyz", "tk": "Turkmen",
        "uk": "Ukrainian", "pl": "Polish", "nl": "Dutch", "id": "Indonesian",
        "ms": "Malay", "ro": "Romanian", "cs": "Czech", "hu": "Hungarian",
    }
    if default_lang == "auto":
        # Try to guess from guest nationality
        nationality = (guest.get("nationality") or "").strip().lower()
        nat_lang_map = {
            "russia": "Russian", "ukraine": "Ukrainian", "belarus": "Russian",
            "germany": "German", "france": "French", "spain": "Spanish",
            "turkey": "Turkish", "china": "Chinese", "japan": "Japanese",
            "korea": "Korean", "arabic": "Arabic", "saudi": "Arabic",
            "uae": "Arabic", "italy": "Italian", "portugal": "Portuguese",
            "brazil": "Portuguese", "uzbekistan": "Uzbek", "kazakhstan": "Kazakh",
            "kyrgyzstan": "Kyrgyz", "azerbaijan": "Azerbaijani", "turkmenistan": "Turkmen",
            "poland": "Polish", "netherlands": "Dutch", "indonesia": "Indonesian",
            "india": "Hindi", "iran": "Persian", "malaysia": "Malay",
            "romania": "Romanian", "czechia": "Czech", "hungary": "Hungarian",
            "uk": "English", "usa": "English",
        }
        guess_lang = nat_lang_map.get(nationality, "English")
        lang_instruction = f"Write the welcome message in {guess_lang} (based on guest nationality)."
    else:
        lang_name = lang_map.get(default_lang, default_lang.upper())
        lang_instruction = f"Write the welcome message in {lang_name}."

    prompt = (
        f"You are '{ai_name}', the AI concierge of {hotel['name']} hotel.\n"
        f"Guest: {first_name} {guest['last_name']}, room {room}, {stay_info}.\n\n"
        f"Hotel info:\n{hotel['info']}\n\n"
        f"Greet the guest by first name as '{ai_name}'. 2-3 sentences, warm and brief. "
        f"Mention a few things you can help with (room service, wifi, check-out time, etc.). "
        f"{lang_instruction}"
    )
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=200,
            messages=[{"role": "user", "content": prompt}]
        )
        welcome_text = resp.content[0].text.strip()
    except Exception as e:
        print(f"[welcome] Claude error: {e}")
        return {"skip": True}

    save_hotel_message(slug, room, "bot", welcome_text)
    return {"message": welcome_text}


# ===== API MESSAGES =====
@app.get("/api/messages")
def api_messages():
    rows = get_messages()
    return [{"room": r[0], "role": r[1], "message": r[2], "created_at": r[3], "priority": r[4], "is_read": r[5]} for r in rows]

@app.post("/api/mark-read")
def mark_read(request: Request):
    if request.cookies.get("manager_auth") != "yes":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    mark_all_read()
    return {"status": "ok"}

@app.get("/api/hotel/{slug}/export")
def hotel_export(slug: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)

    rows = get_hotel_messages(slug)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mesajlar"

    header_fill = PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=12)

    headers = ["Oda", "Kim", "Mesaj", "Öncelik", "Saat", "Okundu"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 10

    for i, row in enumerate(rows, 2):
        room, role, message, created_at, priority, is_read = row
        role_label = {"user": "Misafir", "bot": "AI", "staff": "Personel"}.get(role, role)
        ws.cell(row=i, column=1, value=f"Oda {room}")
        ws.cell(row=i, column=2, value=role_label)
        ws.cell(row=i, column=3, value=message)
        ws.cell(row=i, column=4, value="🔴 Acil" if priority == "urgent" else "🟢 Normal")
        ws.cell(row=i, column=5, value=created_at)
        ws.cell(row=i, column=6, value="✓" if is_read else "○")

        if priority == "urgent":
            for col in range(1, 7):
                ws.cell(row=i, column=col).fill = PatternFill(
                    start_color="2D1515", end_color="2D1515", fill_type="solid"
                )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{hotel['name']}-mesajlar.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/hotel/{slug}/guests/export")
def hotel_guests_export(slug: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)

    guests = get_guests(slug)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Misafirler"

    header_fill = PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=11)
    center = Alignment(horizontal="center")

    headers = ["#", "Ad", "Soyad", "Pasaport", "Uyruk", "Oda", "Giriş", "Çıkış", "Durum", "Notlar", "Kayıt"]
    widths   = [5,   14,    14,      16,         14,       8,      12,      12,       14,      24,       18]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w

    STATUS_MAP = {"pending": "Bekliyor", "checked_in": "İçeride", "checked_out": "Çıktı"}
    checkin_fill  = PatternFill(start_color="0D2010", end_color="0D2010", fill_type="solid")
    checkout_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")

    for row_i, g in enumerate(guests, 2):
        values = [
            row_i - 1,
            g["first_name"], g["last_name"], g["passport"], g["nationality"],
            g["room"], g["check_in"], g["check_out"],
            STATUS_MAP.get(g["status"], g["status"]),
            g["notes"], g["created_at"]
        ]
        fill = checkin_fill if g["status"] == "checked_in" else (checkout_fill if g["status"] == "checked_out" else None)
        for col_i, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{hotel['name']}-misafirler.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@app.get("/api/hotel/{slug}/ratings")
def hotel_ratings(slug: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    return get_recent_ratings(slug)

@app.get("/api/hotel/{slug}/plan")
def hotel_plan(slug: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)
    plan = hotel.get("plan", "trial")
    limit = PLAN_LIMITS.get(plan, PLAN_LIMITS["trial"])
    used = get_monthly_message_count(slug)
    return {
        "plan": plan,
        "limit": limit,
        "used": used,
        "unlimited": limit == -1,
        "percent": 0 if limit == -1 else round(used / limit * 100, 1)
    }


@app.get("/api/hotel/{slug}/stats")
def hotel_stats(slug: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    conn = sqlite3.connect(DATABASE_PATH)

    # --- Message activity (last 7 days) ---
    days = conn.execute("""
        SELECT
            substr(created_at, 1, 10) as day,
            COUNT(*) as total,
            SUM(CASE WHEN priority='urgent' THEN 1 ELSE 0 END) as urgent,
            SUM(CASE WHEN role='user' THEN 1 ELSE 0 END) as guests
        FROM messages
        WHERE hotel_slug=?
        AND created_at >= date('now', '-7 days')
        GROUP BY day
        ORDER BY day ASC
    """, (slug,)).fetchall()

    top_rooms = conn.execute("""
        SELECT room, COUNT(*) as count
        FROM messages
        WHERE hotel_slug=? AND role='user'
        GROUP BY room
        ORDER BY count DESC
        LIMIT 5
    """, (slug,)).fetchall()

    # --- Rating distribution (1-5 stars) ---
    rating_dist_rows = conn.execute("""
        SELECT rating, COUNT(*) as cnt
        FROM ratings WHERE hotel_slug=?
        GROUP BY rating ORDER BY rating
    """, (slug,)).fetchall()
    rating_dist = {str(i): 0 for i in range(1, 6)}
    for row in rating_dist_rows:
        rating_dist[str(row[0])] = row[1]

    # --- Rating trend (avg per day, last 7 days) ---
    rating_trend = conn.execute("""
        SELECT substr(created_at, 1, 10) as day, ROUND(AVG(rating), 2) as avg_r, COUNT(*) as cnt
        FROM ratings WHERE hotel_slug=?
        AND created_at >= date('now', '-7 days')
        GROUP BY day ORDER BY day ASC
    """, (slug,)).fetchall()

    # --- Guest stats ---
    guest_totals = conn.execute("""
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status='checked_in' THEN 1 ELSE 0 END) as checked_in,
            SUM(CASE WHEN status='checked_out' THEN 1 ELSE 0 END) as checked_out
        FROM guests WHERE hotel_slug=?
    """, (slug,)).fetchone()

    # Average stay in days
    avg_stay = conn.execute("""
        SELECT AVG(
            CAST(julianday(check_out) - julianday(check_in) AS REAL)
        )
        FROM guests WHERE hotel_slug=? AND check_out > check_in
    """, (slug,)).fetchone()[0]

    # Nationality breakdown (top 5)
    nat_rows = conn.execute("""
        SELECT nationality, COUNT(*) as cnt
        FROM guests WHERE hotel_slug=? AND nationality != ''
        GROUP BY nationality ORDER BY cnt DESC LIMIT 5
    """, (slug,)).fetchall()

    conn.close()

    avg_rating, rating_count = get_hotel_avg_rating(slug)

    return {
        # Message charts (existing)
        "days": [{"day": r[0], "total": r[1], "urgent": r[2], "guests": r[3]} for r in days],
        "top_rooms": [{"room": r[0], "count": r[1]} for r in top_rooms],
        "avg_rating": avg_rating,
        "rating_count": rating_count,
        # New analytics
        "rating_dist": rating_dist,
        "rating_trend": [{"day": r[0], "avg": r[1], "count": r[2]} for r in rating_trend],
        "guests": {
            "total": guest_totals[0] or 0,
            "checked_in": guest_totals[1] or 0,
            "checked_out": guest_totals[2] or 0,
            "avg_stay_days": round(avg_stay, 1) if avg_stay else None,
        },
        "nationalities": [{"name": r[0], "count": r[1]} for r in nat_rows],
    }

@app.get("/api/hotel/{slug}/analytics")
def hotel_analytics(slug: str, request: Request):
    """Rich analytics for charts: 30-day messages, hourly activity, request categories, funnel."""
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    conn = sqlite3.connect(DATABASE_PATH)

    # ── Messages by day (last 30 days) ──────────────────────────────────────
    rows_30 = conn.execute("""
        SELECT substr(created_at, 1, 10) as day, COUNT(*) as cnt
        FROM messages
        WHERE hotel_slug=? AND role='user'
          AND created_at >= date('now', '-29 days')
        GROUP BY day ORDER BY day ASC
    """, (slug,)).fetchall()
    # Fill missing days with 0
    from datetime import datetime as _dt, timedelta as _td
    today = _dt.utcnow().date()
    day_map = {r[0]: r[1] for r in rows_30}
    messages_by_day = []
    for i in range(29, -1, -1):
        d = str(today - _td(days=i))
        messages_by_day.append({"date": d, "count": day_map.get(d, 0)})

    # ── Messages by hour (last 30 days, guest only) ──────────────────────────
    rows_hour = conn.execute("""
        SELECT CAST(substr(created_at, 12, 2) AS INTEGER) as hour, COUNT(*) as cnt
        FROM messages
        WHERE hotel_slug=? AND role='user'
          AND created_at >= date('now', '-29 days')
        GROUP BY hour ORDER BY hour ASC
    """, (slug,)).fetchall()
    hour_map = {r[0]: r[1] for r in rows_hour}
    messages_by_hour = [{"hour": h, "count": hour_map.get(h, 0)} for h in range(24)]

    # ── Top request categories ───────────────────────────────────────────────
    cat_rows = conn.execute("""
        SELECT category, COUNT(*) as cnt
        FROM requests WHERE hotel_slug=?
        GROUP BY category ORDER BY cnt DESC LIMIT 8
    """, (slug,)).fetchall()
    top_categories = [{"category": r[0], "count": r[1]} for r in cat_rows]

    # ── Rating distribution ──────────────────────────────────────────────────
    rdist_rows = conn.execute("""
        SELECT rating, COUNT(*) as cnt
        FROM ratings WHERE hotel_slug=?
        GROUP BY rating ORDER BY rating ASC
    """, (slug,)).fetchall()
    rating_dist = [0] * 5
    for r in rdist_rows:
        if 1 <= r[0] <= 5:
            rating_dist[r[0] - 1] = r[1]

    # ── Conversion funnel ────────────────────────────────────────────────────
    total_guests = conn.execute(
        "SELECT COUNT(*) FROM guests WHERE hotel_slug=?", (slug,)
    ).fetchone()[0] or 0
    guests_messaged = conn.execute("""
        SELECT COUNT(DISTINCT room) FROM messages
        WHERE hotel_slug=? AND role='user'
    """, (slug,)).fetchone()[0] or 0
    guests_rated = conn.execute(
        "SELECT COUNT(DISTINCT room) FROM ratings WHERE hotel_slug=?", (slug,)
    ).fetchone()[0] or 0

    # ── Avg response time (bot reply latency in minutes) ────────────────────
    # For each user message, find the next bot reply in the same room
    response_rows = conn.execute("""
        SELECT AVG(diff_min) FROM (
            SELECT
              (julianday(b.created_at) - julianday(u.created_at)) * 1440 AS diff_min
            FROM messages u
            JOIN messages b ON b.hotel_slug=u.hotel_slug
                            AND b.room=u.room
                            AND b.role='bot'
                            AND b.id = (
                              SELECT MIN(id) FROM messages
                              WHERE hotel_slug=u.hotel_slug AND room=u.room
                                AND role='bot' AND id > u.id
                            )
            WHERE u.hotel_slug=? AND u.role='user'
              AND diff_min BETWEEN 0 AND 60
        )
    """, (slug,)).fetchone()[0]

    conn.close()
    return {
        "messages_by_day": messages_by_day,
        "messages_by_hour": messages_by_hour,
        "top_categories": top_categories,
        "rating_dist": rating_dist,
        "funnel": {
            "check_ins": total_guests,
            "messaged": guests_messaged,
            "rated": guests_rated,
        },
        "avg_response_min": round(response_rows, 1) if response_rows else None,
    }


@app.get("/api/hotel/{slug}/messages")
def hotel_messages(slug: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    rows = get_hotel_messages(slug)
    return [{"room": r[0], "role": r[1], "message": r[2], "created_at": r[3], "priority": r[4], "is_read": r[5]} for r in rows]

@app.post("/api/hotel/{slug}/mark-read")
def hotel_mark_read_api(slug: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    mark_hotel_read(slug)
    return {"status": "ok"}

@app.get("/api/hotel/{slug}/info")
def hotel_info(slug: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return {
        "name": hotel["name"],
        "info": hotel["info"],
        "telegram_token": hotel.get("telegram_token", ""),
        "telegram_chat_id": hotel.get("telegram_chat_id", ""),
        "booking_url": hotel.get("booking_url", ""),
        "ai_name": hotel.get("ai_name", "AI Asistan"),
        "smtp_host": hotel.get("smtp_host", ""),
        "smtp_port": hotel.get("smtp_port", 587),
        "smtp_user": hotel.get("smtp_user", ""),
        "smtp_from": hotel.get("smtp_from", ""),
        "notify_email": hotel.get("notify_email", ""),
        "default_language": hotel.get("default_language", "auto"),
        "supported_languages": hotel.get("supported_languages", "en,ru,tr,ar,de,fr"),
        "photo_url": hotel.get("photo_url", ""),
        "page_description": hotel.get("page_description", ""),
        "amenities": hotel.get("amenities", ""),
        "rooms_per_floor": hotel.get("rooms_per_floor", 0),
        # smtp_pass intentionally omitted from response
    }

@app.post("/api/hotel/{slug}/update")
def hotel_update(slug: str, data: UpdateRequest, request: Request):
    if not get_auth_role(request, slug)[0]:
        return {"ok": False, "error": "Unauthorized"}
    name = data.name.strip()
    info = data.info.strip()
    if not name or not info:
        return {"ok": False, "error": "Ad ve bilgiler gerekli"}
    update_hotel(slug, name, info,
                 data.password.strip() if data.password else None,
                 data.tg_token.strip(), data.tg_chat.strip(),
                 booking_url=data.booking_url.strip(),
                 ai_name=data.ai_name.strip() if data.ai_name else None,
                 smtp_host=data.smtp_host,
                 smtp_port=data.smtp_port,
                 smtp_user=data.smtp_user,
                 smtp_pass=data.smtp_pass if data.smtp_pass else None,
                 smtp_from=data.smtp_from,
                 notify_email=data.notify_email,
                 default_language=data.default_language,
                 supported_languages=data.supported_languages,
                 photo_url=data.photo_url,
                 page_description=data.page_description,
                 amenities=data.amenities,
                 rooms_per_floor=data.rooms_per_floor)
    return {"ok": True}

# ===== EDIT =====
@app.get("/hotel/{slug}/edit", response_class=HTMLResponse)
def hotel_edit(slug: str, request: Request):
    hotel = get_hotel(slug)
    if not hotel:
        return HTMLResponse("❌ Otel bulunamadı", status_code=404)
    if not get_auth_role(request, slug)[0]:
        return RedirectResponse(f"/hotel/{slug}/login")
    return get_edit_html()

# ===== QR CODES =====
def _make_qr(url: str) -> bytes:
    """Generate QR code PNG bytes for a given URL."""
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#C9A84C", back_color="#0a0a0a")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf.getvalue()

@app.get("/hotel/{slug}/qr/{room}")
def hotel_qr(slug: str, room: str, request: Request):
    base = str(request.base_url).rstrip("/")
    url = f"{base}/hotel/{slug}?room={room}"
    return Response(content=_make_qr(url), media_type="image/png")

@app.get("/hotel/{slug}/qr-checkin")
def hotel_qr_checkin(slug: str, request: Request):
    """Dedicated QR for the self check-in page."""
    base = str(request.base_url).rstrip("/")
    url = f"{base}/hotel/{slug}/checkin"
    return Response(content=_make_qr(url), media_type="image/png")

@app.get("/hotel/{slug}/qrcodes", response_class=HTMLResponse)
def hotel_qrcodes(slug: str, request: Request):
    hotel = get_hotel(slug)
    if not hotel:
        return HTMLResponse("❌ Otel bulunamadı", status_code=404)

    room_start = hotel.get("room_start", 101)
    room_count = hotel.get("room_count", 30)
    rooms_per_floor = hotel.get("rooms_per_floor", 0)
    rooms = generate_room_numbers(room_count, room_start, rooms_per_floor)

    # Room cards
    room_cards = "".join([f"""
        <div class="card">
            <img src="/hotel/{slug}/qr/{room}" loading="lazy">
            <div class="room">🚪 {room}</div>
        </div>""" for room in rooms])

    # Zone cards — data-zone key used for JS translation
    zones = [
        ("reception",  "🛎️", "Resepsiyon"),
        ("restaurant", "🍽️", "Restoran"),
        ("pool",       "🏊", "Havuz"),
        ("spa",        "💆", "Spa"),
    ]
    zone_cards = "".join([f"""
        <div class="card zone-card">
            <img src="/hotel/{slug}/qr/{zone_id}" loading="lazy">
            <div class="room" data-zone="{zone_id}">{icon} {label}</div>
        </div>""" for zone_id, icon, label in zones])

    return f"""<!DOCTYPE html>
    <html><head><meta charset="utf-8"><title>QR — {hotel['name']}</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
    <style>
        *{{margin:0;padding:0;box-sizing:border-box}}
        body{{background:#0a0a0a;color:white;font-family:'Inter',sans-serif;padding:40px}}
        h1{{color:#C9A84C;font-size:28px;margin-bottom:4px}}
        .subtitle{{color:#555;margin-bottom:32px;font-size:14px}}
        h2{{color:#888;font-size:13px;letter-spacing:2px;text-transform:uppercase;margin:32px 0 16px;padding-bottom:10px;border-bottom:1px solid #1a1a1a}}
        .grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:16px}}
        .card{{background:#111;border-radius:12px;padding:20px;text-align:center;border:1px solid #1f1f1f}}
        .zone-card{{border-color:rgba(201,168,76,0.2);background:#13110a}}
        .card img{{width:140px;height:140px;border-radius:8px}}
        .room{{font-size:14px;font-weight:700;color:#C9A84C;margin-top:10px}}
        .btns{{display:flex;gap:12px;margin-bottom:32px;flex-wrap:wrap}}
        .btn{{background:#C9A84C;color:#000;border:none;padding:11px 28px;border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;font-family:inherit}}
        .btn-dark{{background:#1a1a1a;color:#fff;border:1px solid #2a2a2a}}
        @media print{{.btns{{display:none}}}}
    </style></head>
    <body>
        <h1 id="qrTitle">🏨 {hotel['name']} — QR Kodlar</h1>
        <p class="subtitle" id="qrSubtitle">Her odaya yapıştırın. Misafir tarar → sohbete girer. ({room_count} oda)</p>
        <div class="btns">
            <button class="btn" id="qrPrint" onclick="window.print()">🖨️ Yazdır</button>
            <a class="btn btn-dark" id="qrBack" href="/hotel/{slug}/dashboard">← Panele Dön</a>
        </div>

        <h2 id="qrH2Checkin">✅ Self Check-in QR</h2>
        <div style="background:#13110a;border:1px solid rgba(201,168,76,0.25);border-radius:16px;padding:28px;display:flex;align-items:center;gap:28px;margin-bottom:8px;max-width:420px">
            <img src="/hotel/{slug}/qr-checkin" style="width:160px;height:160px;border-radius:10px;flex-shrink:0">
            <div>
                <div style="font-size:18px;font-weight:700;color:#C9A84C;margin-bottom:8px" id="qrCheckinTitle">Dijital Check-in</div>
                <div style="font-size:13px;color:#666;line-height:1.7" id="qrCheckinDesc">
                    Resepsiyona yapıştırın.<br>
                    Misafir tarar → formu doldurur → siz panelde görürsünüz.
                </div>
                <div style="margin-top:14px;font-size:12px;color:#333;background:#0a0a0a;padding:8px 12px;border-radius:8px;word-break:break-all">
                    /hotel/{slug}/checkin
                </div>
            </div>
        </div>

        <h2 id="qrH2Zones">📱 Bölge QR Kodları</h2>
        <div class="grid">{zone_cards}</div>

        <h2 id="qrH2Rooms">🚪 Oda QR Kodları ({rooms[0]} – {rooms[-1]})</h2>
        <div class="grid">{room_cards}</div>

    <script>
    (function() {{
        const lang = localStorage.getItem('ss_lang') || 'en';
        const name = '{hotel['name']}';
        const roomCount = {room_count};
        const roomRange = '{rooms[0]} – {rooms[-1]}';
        const L = {{
            en: {{
                title: '🏨 ' + name + ' — QR Codes',
                subtitle: 'Print and place on each door. Guest scans → enters chat. (' + roomCount + ' rooms)',
                print: '🖨️ Print', back: '← Back to Dashboard',
                h2Checkin: '✅ Self Check-in QR', checkinTitle: 'Digital Check-in',
                checkinDesc: 'Place at reception.<br>Guest scans → fills form → you see it in the dashboard.',
                h2Zones: '📱 Zone QR Codes', h2Rooms: '🚪 Room QR Codes (' + roomRange + ')',
                zones: {{reception:'🛎️ Reception', restaurant:'🍽️ Restaurant', pool:'🏊 Pool', spa:'💆 Spa'}},
            }},
            ru: {{
                title: '🏨 ' + name + ' — QR-коды',
                subtitle: 'Разместите в каждом номере. Гость сканирует → входит в чат. (' + roomCount + ' номеров)',
                print: '🖨️ Печать', back: '← Назад',
                h2Checkin: '✅ QR для самостоятельного заезда', checkinTitle: 'Цифровой чек-ин',
                checkinDesc: 'Разместите на ресепшен.<br>Гость сканирует → заполняет форму → вы видите в панели.',
                h2Zones: '📱 QR-коды зон', h2Rooms: '🚪 QR-коды номеров (' + roomRange + ')',
                zones: {{reception:'🛎️ Ресепшен', restaurant:'🍽️ Ресторан', pool:'🏊 Бассейн', spa:'💆 Спа'}},
            }},
            tr: {{
                title: '🏨 ' + name + ' — QR Kodlar',
                subtitle: 'Her odaya yapıştırın. Misafir tarar → sohbete girer. (' + roomCount + ' oda)',
                print: '🖨️ Yazdır', back: '← Panele Dön',
                h2Checkin: '✅ Self Check-in QR', checkinTitle: 'Dijital Check-in',
                checkinDesc: 'Resepsiyona yapıştırın.<br>Misafir tarar → formu doldurur → siz panelde görürsünüz.',
                h2Zones: '📱 Bölge QR Kodları', h2Rooms: '🚪 Oda QR Kodları (' + roomRange + ')',
                zones: {{reception:'🛎️ Resepsiyon', restaurant:'🍽️ Restoran', pool:'🏊 Havuz', spa:'💆 Spa'}},
            }},
            uz: {{
                title: "🏨 " + name + " — QR kodlar",
                subtitle: "Har bir xonaga yopishtirib qo'ying. Mehmon skaner qiladi → chatga kiradi. (" + roomCount + " xona)",
                print: "🖨️ Chop etish", back: "← Panelga",
                h2Checkin: "✅ O'z-o'ziga kirish QR", checkinTitle: "Raqamli Check-in",
                checkinDesc: "Resepsiyanga yopishtirib qo'ying.<br>Mehmon skaner qiladi → forma to'ldiradi → siz panelda ko'rasiz.",
                h2Zones: "📱 Zona QR kodlar", h2Rooms: "🚪 Xona QR kodlar (" + roomRange + ")",
                zones: {{reception:"🛎️ Resepsiya", restaurant:"🍽️ Restoran", pool:"🏊 Hovuz", spa:"💆 Spa"}},
            }},
        }};
        const T = L[lang] || L.en;
        document.getElementById('qrTitle').textContent = T.title;
        document.getElementById('qrSubtitle').textContent = T.subtitle;
        document.getElementById('qrPrint').textContent = T.print;
        document.getElementById('qrBack').textContent = T.back;
        document.getElementById('qrH2Checkin').textContent = T.h2Checkin;
        document.getElementById('qrCheckinTitle').textContent = T.checkinTitle;
        document.getElementById('qrCheckinDesc').innerHTML = T.checkinDesc;
        document.getElementById('qrH2Zones').textContent = T.h2Zones;
        document.getElementById('qrH2Rooms').textContent = T.h2Rooms;
        document.querySelectorAll('[data-zone]').forEach(el => {{
            const key = el.getAttribute('data-zone');
            if (T.zones[key]) el.textContent = T.zones[key];
        }});
    }})();
    </script>
    </body></html>"""

@app.get("/qr/{room}")
def get_qr(room: str, request: Request):
    base = str(request.base_url).rstrip("/")
    url = f"{base}?room={room}"
    return Response(content=_make_qr(url), media_type="image/png")

@app.get("/qrcodes", response_class=HTMLResponse)
def qrcodes():
    rooms = list(range(101, 111)) + list(range(201, 211)) + list(range(301, 311))
    cards = "".join([f"""
        <div class="card">
            <img src="/qr/{room}">
            <div class="room">🚪 {room}</div>
        </div>""" for room in rooms])
    return f"""<!DOCTYPE html>
    <html><head><meta charset="utf-8"><title>QR Kodlar</title>
    <style>
        *{{margin:0;padding:0;box-sizing:border-box}}
        body{{background:#0a0a0a;color:white;font-family:sans-serif;padding:40px}}
        h1{{color:#C9A84C;font-size:28px;margin-bottom:32px}}
        .grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:20px}}
        .card{{background:#111;border-radius:12px;padding:24px;text-align:center;border:1px solid #222}}
        .card img{{width:140px;height:140px;border-radius:8px}}
        .room{{font-size:16px;font-weight:700;color:#C9A84C;margin-top:12px}}
        .btn{{background:#C9A84C;color:#000;border:none;padding:12px 32px;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;margin-bottom:32px}}
    </style></head>
    <body>
        <h1>🏨 QR Kodlar</h1>
        <button class="btn" onclick="window.print()">🖨️ Yazdır</button>
        <div class="grid">{cards}</div>
    </body></html>"""

# ===== ROOM NOTES =====

class RoomNoteRequest(BaseModel):
    note: str
    author: str = "Персонал"


@app.get("/api/hotel/{slug}/room/{room}/notes")
def list_room_notes(slug: str, room: str, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    return get_room_notes(slug, room)


@app.post("/api/hotel/{slug}/room/{room}/notes")
def add_room_note(slug: str, room: str, body: RoomNoteRequest, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    note = body.note.strip()
    if not note:
        return JSONResponse({"error": "Empty note"}, status_code=400)
    note_id = save_room_note(slug, room, note, body.author)
    return {"ok": True, "id": note_id}


@app.delete("/api/hotel/{slug}/room/{room}/notes/{note_id}")
def remove_room_note(slug: str, room: str, note_id: int, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    ok = delete_room_note(note_id, slug)
    if not ok:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return {"ok": True}


# ===== GUEST REQUEST TRACKER =====

class RequestStatusUpdate(BaseModel):
    status: str  # pending | in_progress | resolved

class ServiceRequest(BaseModel):
    name: str
    description: str = ""
    category: str = "general"
    price: float = 0
    currency: str = "USD"
    icon: str = "🛎️"
    is_active: bool = True
    sort_order: int = 0

class ManualRequestCreate(BaseModel):
    room: str
    category: str = "general"
    message: str
    guest_name: str = ""


@app.post("/api/hotel/{slug}/requests")
async def create_request(slug: str, body: ManualRequestCreate, request: Request):
    # Accept manager auth, staff auth, or hotel_slug cookie (dashboard uses all three)
    hotel_slug_cookie = request.cookies.get("hotel_slug")
    authorized, _ = get_auth_role(request, slug)
    if not authorized and hotel_slug_cookie != slug:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    if not body.room.strip() or not body.message.strip():
        return JSONResponse({"error": "Oda ve talep metni gerekli"}, status_code=400)
    valid_cats = {"room_service", "maintenance", "housekeeping", "general"}
    cat = body.category if body.category in valid_cats else "general"
    req_id = save_request(slug, body.room.strip(), body.guest_name.strip(), cat, body.message.strip())
    return {"ok": True, "id": req_id, "pending_count": get_pending_requests_count(slug)}


@app.get("/api/hotel/{slug}/requests")
async def list_requests(slug: str, request: Request, status: str = ""):
    hotel_slug_cookie = request.cookies.get("hotel_slug")
    authorized, _ = get_auth_role(request, slug)
    if not authorized and hotel_slug_cookie != slug:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    reqs = get_requests(slug, status if status else None)
    badge = get_pending_requests_count(slug)
    return {"requests": reqs, "pending_count": badge}


@app.patch("/api/hotel/{slug}/requests/{req_id}")
async def patch_request(slug: str, req_id: int, body: RequestStatusUpdate, request: Request):
    hotel_slug_cookie = request.cookies.get("hotel_slug")
    authorized, _ = get_auth_role(request, slug)
    if not authorized and hotel_slug_cookie != slug:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    valid = {"pending", "in_progress", "resolved"}
    if body.status not in valid:
        return JSONResponse({"error": "Invalid status"}, status_code=400)
    ok = update_request_status(req_id, slug, body.status)
    if not ok:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return {"ok": True, "pending_count": get_pending_requests_count(slug)}


@app.delete("/api/hotel/{slug}/requests/{req_id}")
async def remove_request(slug: str, req_id: int, request: Request):
    hotel_slug_cookie = request.cookies.get("hotel_slug")
    authorized, _ = get_auth_role(request, slug)
    if not authorized and hotel_slug_cookie != slug:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    ok = delete_request(req_id, slug)
    if not ok:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return {"ok": True}


# ===== SERVICES CATALOG =====

@app.get("/api/hotel/{slug}/services")
def list_services(slug: str, active_only: bool = False):
    """Public endpoint — no auth required. Returns hotel's service catalog."""
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Hotel not found"}, status_code=404)
    return get_services(slug, active_only=active_only)


@app.post("/api/hotel/{slug}/services")
def add_service(slug: str, data: ServiceRequest, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    name = data.name.strip()
    if not name:
        return JSONResponse({"error": "Name is required"}, status_code=400)
    svc_id = create_service(
        slug, name, data.description, data.category,
        data.price, data.currency, data.icon, data.sort_order
    )
    return {"ok": True, "id": svc_id}


@app.put("/api/hotel/{slug}/services/{svc_id}")
def edit_service(slug: str, svc_id: int, data: ServiceRequest, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    ok = update_service(
        svc_id, slug,
        name=data.name, description=data.description, category=data.category,
        price=data.price, currency=data.currency, icon=data.icon,
        is_active=data.is_active, sort_order=data.sort_order
    )
    if not ok:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return {"ok": True}


@app.patch("/api/hotel/{slug}/services/{svc_id}/toggle")
def toggle_service(slug: str, svc_id: int, request: Request):
    """Toggle is_active without a full update."""
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    conn = sqlite3.connect(DATABASE_PATH)
    row = conn.execute(
        "SELECT is_active FROM services WHERE id=? AND hotel_slug=?", (svc_id, slug)
    ).fetchone()
    if not row:
        conn.close()
        return JSONResponse({"error": "Not found"}, status_code=404)
    new_state = 0 if row[0] else 1
    conn.execute("UPDATE services SET is_active=? WHERE id=? AND hotel_slug=?",
                 (new_state, svc_id, slug))
    conn.commit()
    conn.close()
    return {"ok": True, "is_active": bool(new_state)}


@app.delete("/api/hotel/{slug}/services/{svc_id}")
def remove_service(slug: str, svc_id: int, request: Request):
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    ok = delete_service(svc_id, slug)
    if not ok:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return {"ok": True}


# ===== STAFF CHAT =====

class StaffChatMessage(BaseModel):
    message: str

@app.get("/api/hotel/{slug}/staff-chat/{channel}")
def staff_chat_history(slug: str, channel: str, request: Request):
    """Returns last 60 messages in a staff channel. Staff or manager auth required."""
    ok, role = get_auth_role(request, slug)
    if not ok:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    if channel not in STAFF_CHANNELS:
        return JSONResponse({"error": "Invalid channel"}, status_code=400)
    msgs = get_staff_msgs(slug, channel)
    return {"messages": msgs}


@app.post("/api/hotel/{slug}/staff-chat/{channel}")
def staff_chat_send(slug: str, channel: str, body: StaffChatMessage, request: Request):
    """Send a message to a staff channel. Staff or manager auth required."""
    ok, role = get_auth_role(request, slug)
    if not ok:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    if channel not in STAFF_CHANNELS:
        return JSONResponse({"error": "Invalid channel"}, status_code=400)
    text = body.message.strip()
    if not text:
        return JSONResponse({"error": "Message is empty"}, status_code=400)

    # Resolve sender name from cookie
    if role == "manager":
        sender = "Менеджер"
        role_label = "manager"
    else:
        staff_id_str = request.cookies.get(f"staff_{slug}", "")
        try:
            staff = get_staff_by_id(slug, int(staff_id_str))
            sender = staff["name"] if staff else "Персонал"
            role_label = staff["role"] if staff else role
        except Exception:
            sender = "Персонал"
            role_label = role

    msg_id = save_staff_msg(slug, channel, sender, role_label, text)
    return {"ok": True, "id": msg_id}


@app.get("/api/hotel/{slug}/staff-chat/{channel}/poll")
def staff_chat_poll(slug: str, channel: str, since_id: int = 0, request: Request = None):
    """Polling endpoint — returns messages after since_id. Staff or manager auth required."""
    ok, role = get_auth_role(request, slug)
    if not ok:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    if channel not in STAFF_CHANNELS:
        return JSONResponse({"error": "Invalid channel"}, status_code=400)
    msgs = get_staff_msgs_since(slug, channel, since_id)
    return {"messages": msgs}


# ===== CRON JOBS =====
@app.get("/api/cron/daily-digest")
async def cron_daily_digest(secret: str = "", request: Request = None):
    """
    Morning digest: for every hotel with Telegram configured, sends a summary
    and individual check-out reminders.

    Call via Railway cron or external scheduler:
      GET https://your-domain.com/api/cron/daily-digest?secret=YOUR_CRON_SECRET

    railway.toml cron example:
      [[cron]]
      schedule = "0 7 * * *"   # 07:00 UTC every day
      command  = "curl -s $RAILWAY_PUBLIC_DOMAIN/api/cron/daily-digest?secret=$CRON_SECRET"
    """
    if not CRON_SECRET or secret != CRON_SECRET:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    hotels = get_all_hotels()
    sent = 0
    emails_sent = 0

    for h in hotels:
        hotel = get_hotel(h["slug"])
        if not hotel:
            continue

        has_telegram = hotel.get("telegram_token") and hotel.get("telegram_chat_id")
        has_email = hotel.get("notify_email") and hotel.get("smtp_host")
        if not has_telegram and not has_email:
            continue

        d = get_daily_digest_data(h["slug"])
        today_str = date.today().strftime("%d.%m.%Y")
        rating_str = f"{d['avg_rating']} ★ ({d['rating_count']} yorum)" if d["avg_rating"] else "— yorum yok"

        # ─── Telegram digest ───────────────────────────────────────
        if has_telegram:
            lines = [
                f"☀️ <b>Günlük Özet — {today_str}</b>",
                f"🏨 {hotel['name']}",
                "",
                f"🛎️ Aktif misafir: <b>{d['active_guests']}</b>",
                f"📥 Bugün check-in: <b>{d['checkins_today']}</b>",
                f"📤 Bugün check-out: <b>{len(d['checkouts_today'])}</b>",
                f"📬 Okunmamış mesaj: <b>{d['unread_messages']}</b>",
                f"⭐ Ortalama puan: <b>{rating_str}</b>",
            ]
            await send_telegram(
                "\n".join(lines),
                token=hotel["telegram_token"],
                chat_id=hotel["telegram_chat_id"]
            )
            sent += 1

            # Individual checkout reminders via Telegram
            for guest in d["checkouts_today"]:
                reminder = (
                    f"⏰ <b>Check-out Hatırlatması</b>\n"
                    f"👤 {guest['name']}\n"
                    f"🚪 Oda: {guest['room']}\n"
                    f"📅 Bugün çıkış yapacak — gerekli düzenlemeleri yapın."
                )
                await send_telegram(
                    reminder,
                    token=hotel["telegram_token"],
                    chat_id=hotel["telegram_chat_id"]
                )

        # ─── Email digest ──────────────────────────────────────────
        if has_email:
            checkouts_rows = "".join(
                f"<tr><td style='padding:6px 12px'>{g['room']}</td><td style='padding:6px 12px'>{g['name']}</td></tr>"
                for g in d["checkouts_today"]
            ) or "<tr><td colspan='2' style='padding:6px 12px;color:#888'>Bugün çıkış yok</td></tr>"

            email_html = f"""
<div style="font-family:sans-serif;max-width:640px;margin:0 auto">
  <div style="background:#0a0a0a;color:#C9A84C;padding:20px 24px;border-radius:8px 8px 0 0">
    <h2 style="margin:0">☀️ Günlük Özet — {today_str}</h2>
    <p style="margin:4px 0 0;color:#aaa;font-size:14px">{hotel['name']}</p>
  </div>
  <div style="border:1px solid #e0e0e0;border-top:none;padding:24px;border-radius:0 0 8px 8px">
    <table style="width:100%;border-collapse:collapse;margin-bottom:20px">
      <tr style="background:#f5f5f5">
        <td style="padding:10px 14px;font-weight:bold">🛎️ Aktif misafir</td>
        <td style="padding:10px 14px;font-size:20px;font-weight:bold;color:#C9A84C">{d['active_guests']}</td>
      </tr>
      <tr>
        <td style="padding:10px 14px;font-weight:bold">📥 Bugün check-in</td>
        <td style="padding:10px 14px;font-size:20px;font-weight:bold;color:#27ae60">{d['checkins_today']}</td>
      </tr>
      <tr style="background:#f5f5f5">
        <td style="padding:10px 14px;font-weight:bold">📤 Bugün check-out</td>
        <td style="padding:10px 14px;font-size:20px;font-weight:bold;color:#e74c3c">{len(d['checkouts_today'])}</td>
      </tr>
      <tr>
        <td style="padding:10px 14px;font-weight:bold">📬 Okunmamış mesaj</td>
        <td style="padding:10px 14px;font-size:20px;font-weight:bold">{d['unread_messages']}</td>
      </tr>
      <tr style="background:#f5f5f5">
        <td style="padding:10px 14px;font-weight:bold">⭐ Ortalama puan</td>
        <td style="padding:10px 14px">{rating_str}</td>
      </tr>
    </table>

    {'<h3 style="margin-bottom:8px">📤 Bugün Check-out Yapacaklar</h3><table style="width:100%;border-collapse:collapse;border:1px solid #e0e0e0;border-radius:6px"><thead><tr style="background:#f5f5f5"><th style="padding:8px 12px;text-align:left">Oda</th><th style="padding:8px 12px;text-align:left">Misafir</th></tr></thead><tbody>' + checkouts_rows + '</tbody></table>' if d['checkouts_today'] else ''}

    <p style="color:#aaa;font-size:12px;margin-top:20px">SmartStay AI tarafından gönderildi</p>
  </div>
</div>"""
            ok = await send_email(hotel, f"☀️ Günlük Özet — {hotel['name']} ({today_str})", email_html)
            if ok:
                emails_sent += 1

    # ─── Auto-checkout overdue guests (all hotels) ─────────────────────────
    total_checked_out = 0
    for h in hotels:
        count = auto_checkout_overdue(h["slug"])
        if count:
            total_checked_out += count
            hotel = get_hotel(h["slug"])
            # Notify via Telegram if configured
            if hotel and hotel.get("telegram_token") and hotel.get("telegram_chat_id"):
                async def _notify_checkout(ht=hotel, n=count):
                    await send_telegram(
                        f"🔄 <b>Otomatik Check-out</b> — {ht['name']}\n"
                        f"{n} misafir check_out tarihi geçtiği için otomatik olarak çıkış yapıldı.",
                        token=ht["telegram_token"], chat_id=ht["telegram_chat_id"]
                    )
                asyncio.create_task(_notify_checkout())

    return {"ok": True, "hotels_notified": sent, "emails_sent": emails_sent, "auto_checkouts": total_checked_out}


# ===== TELEGRAM 2-WAY WEBHOOK =====
@app.get("/api/hotel/{slug}/telegram/set-webhook")
async def telegram_set_webhook(slug: str, request: Request):
    """
    One-click webhook registration. Call this after setting up the bot token.
    GET /api/hotel/{slug}/telegram/set-webhook  (requires dashboard auth)
    """
    if not get_auth_role(request, slug)[0]:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)
    token = hotel.get("telegram_token", "").strip()
    if not token:
        return JSONResponse({"error": "Telegram token not configured. Set it in Settings first."}, status_code=400)

    base = str(request.base_url).rstrip("/")
    webhook_url = f"{base}/api/telegram/webhook/{slug}/{token}"
    result = await set_webhook(token, webhook_url)
    return {"ok": result.get("ok", False), "webhook_url": webhook_url, "telegram": result}


@app.post("/api/telegram/webhook/{slug}/{token}")
async def telegram_webhook(slug: str, token: str, request: Request):
    """
    Receives updates from Telegram. When staff replies to a guest message,
    the reply is routed back to the correct room in the guest chat.
    """
    hotel = get_hotel(slug)
    if not hotel or hotel.get("telegram_token", "").strip() != token:
        # Token mismatch — silently return 200 (don't expose errors to bots)
        return {"ok": True}

    try:
        update = await request.json()
    except Exception:
        return {"ok": True}

    message = update.get("message", {})
    reply_to = message.get("reply_to_message", {})
    text = message.get("text", "").strip()

    if not text or not reply_to:
        return {"ok": True}

    # Look up which room this reply belongs to
    original_msg_id = reply_to.get("message_id")
    if not original_msg_id:
        return {"ok": True}

    room = get_room_by_telegram_msg_id(slug, original_msg_id)
    if not room:
        return {"ok": True}  # unknown mapping — ignore

    # Save as staff message → appears in guest chat via polling
    save_staff_message(slug, room, text)
    print(f"[telegram 2-way] {slug}/room/{room} ← staff: {text[:60]}")
    return {"ok": True}


# ===== ADMIN =====
@app.get("/admin/login", response_class=HTMLResponse)
def admin_login_page():
    return get_admin_login_html()

@app.post("/api/admin/login")
@limiter.limit("5/minute")
def api_admin_login(data: AdminLoginRequest, request: Request, response: Response):
    if data.password == ADMIN_PASSWORD:
        response.set_cookie("admin_auth", "yes", max_age=86400,
                            httponly=True, secure=SECURE_COOKIES, samesite="strict")
        return {"ok": True}
    return {"ok": False}

@app.get("/admin", response_class=HTMLResponse)
def admin_panel(request: Request):
    if request.cookies.get("admin_auth") != "yes":
        return RedirectResponse("/admin/login")
    return get_admin_html()

@app.get("/admin/logout")
def admin_logout(response: Response):
    response.delete_cookie("admin_auth")
    return RedirectResponse("/admin/login")

class PlanUpdateRequest(BaseModel):
    plan: str

@app.patch("/api/admin/hotel/{slug}/plan")
def api_admin_update_plan(slug: str, data: PlanUpdateRequest, request: Request):
    if request.cookies.get("admin_auth") != "yes":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    if data.plan not in PLAN_LIMITS:
        return JSONResponse({"error": f"Invalid plan. Choose: {list(PLAN_LIMITS)}"}, status_code=400)
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)
    conn = sqlite3.connect(DATABASE_PATH)
    conn.execute("UPDATE hotels SET plan=? WHERE slug=?", (data.plan, slug))
    conn.commit()
    conn.close()
    return {"ok": True, "slug": slug, "plan": data.plan}

class PasswordResetRequest(BaseModel):
    password: str

@app.patch("/api/admin/hotel/{slug}/password")
def api_admin_reset_password(slug: str, data: PasswordResetRequest, request: Request):
    if request.cookies.get("admin_auth") != "yes":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    pw = data.password.strip()
    if len(pw) < 6:
        return JSONResponse({"error": "Şifre en az 6 karakter olmalı"}, status_code=400)
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)
    from database import hash_password
    new_hash = hash_password(pw)
    conn = sqlite3.connect(DATABASE_PATH)
    conn.execute("UPDATE hotels SET password=? WHERE slug=?", (new_hash, slug))
    conn.commit()
    conn.close()
    return {"ok": True, "slug": slug}


@app.delete("/api/admin/hotel/{slug}")
def api_admin_delete_hotel(slug: str, request: Request):
    if request.cookies.get("admin_auth") != "yes":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    delete_hotel(slug)
    return {"ok": True}

@app.get("/api/admin/hotels")
def api_admin_hotels(request: Request):
    if request.cookies.get("admin_auth") != "yes":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    hotels = get_all_hotels()
    conn = sqlite3.connect(DATABASE_PATH)
    total_messages = conn.execute("SELECT COUNT(*) FROM messages").fetchone()[0]
    total_guests = conn.execute("SELECT COUNT(*) FROM guests").fetchone()[0]

    for hotel in hotels:
        slug = hotel["slug"]
        hotel["message_count"] = conn.execute(
            "SELECT COUNT(*) FROM messages WHERE hotel_slug=?", (slug,)
        ).fetchone()[0]
        row = conn.execute(
            "SELECT telegram_token, plan FROM hotels WHERE slug=?", (slug,)
        ).fetchone()
        hotel["telegram_token"] = row[0] or "" if row else ""
        hotel["plan"] = row[1] or "trial" if row else "trial"
        hotel["active_guests"] = conn.execute(
            "SELECT COUNT(*) FROM guests WHERE hotel_slug=? AND status='checked_in'", (slug,)
        ).fetchone()[0]
        hotel["total_guests"] = conn.execute(
            "SELECT COUNT(*) FROM guests WHERE hotel_slug=?", (slug,)
        ).fetchone()[0]
        avg_row = conn.execute(
            "SELECT ROUND(AVG(rating),1), COUNT(*) FROM ratings WHERE hotel_slug=?", (slug,)
        ).fetchone()
        hotel["avg_rating"] = avg_row[0]
        hotel["rating_count"] = avg_row[1]

    conn.close()

    # Revenue projection by plan prices (USD/month)
    PLAN_PRICES = {"trial": 0, "starter": 299, "pro": 599, "premium": 999}
    monthly_revenue = sum(PLAN_PRICES.get(h["plan"], 0) for h in hotels)

    return {
        "hotels": hotels,
        "total_messages": total_messages,
        "total_guests": total_guests,
        "monthly_revenue": monthly_revenue,
    }


# ===== OWNER (MULTI-HOTEL) =====

def get_auth_owner(request: Request) -> dict | None:
    """Return owner dict if owner_session cookie is valid, else None."""
    owner_id_str = request.cookies.get("owner_session", "")
    if not owner_id_str:
        return None
    try:
        return get_owner_by_id(int(owner_id_str))
    except Exception:
        return None

class OwnerLoginRequest(BaseModel):
    email: str
    password: str

class CreateOwnerRequest(BaseModel):
    name: str
    email: str
    password: str

class AssignHotelRequest(BaseModel):
    hotel_slug: str

@app.get("/owner/login", response_class=HTMLResponse)
def owner_login_page():
    return get_owner_login_html()

@app.post("/api/owner/login")
def api_owner_login(data: OwnerLoginRequest, response: Response):
    owner = get_owner_by_email(data.email)
    if not owner or not verify_password(data.password, owner["password_hash"]):
        return JSONResponse({"ok": False, "error": "Неверный email или пароль"}, status_code=401)
    response.set_cookie("owner_session", str(owner["id"]), max_age=86400 * 30,
                        httponly=True, secure=SECURE_COOKIES, samesite="strict")
    return {"ok": True}

@app.get("/owner/dashboard", response_class=HTMLResponse)
def owner_dashboard(request: Request):
    owner = get_auth_owner(request)
    if not owner:
        return RedirectResponse("/owner/login")
    hotels = get_owner_hotels(owner["id"])
    return get_owner_dashboard_html(owner, hotels)

@app.get("/owner/logout")
def owner_logout(response: Response):
    response.delete_cookie("owner_session")
    return RedirectResponse("/owner/login")

# Owner impersonate hotel — set manager cookie and redirect to dashboard
@app.get("/owner/enter/{slug}")
def owner_enter_hotel(slug: str, request: Request, response: Response):
    owner = get_auth_owner(request)
    if not owner:
        return RedirectResponse("/owner/login")
    # Verify this owner actually has access to this hotel
    hotel_owner_ids = get_hotel_owner_ids(slug)
    if owner["id"] not in hotel_owner_ids:
        return JSONResponse({"error": "Нет доступа к этому отелю"}, status_code=403)
    response.set_cookie(f"auth_{slug}", "yes", max_age=86400,
                        httponly=True, secure=SECURE_COOKIES, samesite="strict")
    return RedirectResponse(f"/hotel/{slug}/dashboard")

@app.get("/api/owner/hotels")
def api_owner_hotels(request: Request):
    owner = get_auth_owner(request)
    if not owner:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    hotels = get_owner_hotels(owner["id"])
    conn = sqlite3.connect(DATABASE_PATH)
    for h in hotels:
        slug = h["slug"]
        h["active_guests"] = conn.execute(
            "SELECT COUNT(*) FROM guests WHERE hotel_slug=? AND status='checked_in'", (slug,)
        ).fetchone()[0]
        h["unread_messages"] = conn.execute(
            "SELECT COUNT(*) FROM messages WHERE hotel_slug=? AND is_read=0", (slug,)
        ).fetchone()[0]
        avg_row = conn.execute(
            "SELECT ROUND(AVG(rating),1) FROM ratings WHERE hotel_slug=?", (slug,)
        ).fetchone()
        h["avg_rating"] = avg_row[0] or "—"
    conn.close()
    return {"owner": {"name": owner["name"], "email": owner["email"]}, "hotels": hotels}

# ===== ADMIN — OWNER MANAGEMENT =====

@app.post("/api/admin/owners")
def api_admin_create_owner(data: CreateOwnerRequest, request: Request):
    if request.cookies.get("admin_auth") != "yes":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    owner_id = create_owner(data.name, data.email, data.password)
    if owner_id is None:
        return JSONResponse({"error": "Email уже используется"}, status_code=400)
    return {"ok": True, "owner_id": owner_id}

@app.get("/api/admin/owners")
def api_admin_get_owners(request: Request):
    if request.cookies.get("admin_auth") != "yes":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    return {"owners": get_all_owners()}

@app.post("/api/admin/owners/{owner_id}/hotels")
def api_admin_assign_hotel(owner_id: int, data: AssignHotelRequest, request: Request):
    if request.cookies.get("admin_auth") != "yes":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    assign_hotel_to_owner(owner_id, data.hotel_slug)
    return {"ok": True}

@app.delete("/api/admin/owners/{owner_id}/hotels/{slug}")
def api_admin_remove_hotel(owner_id: int, slug: str, request: Request):
    if request.cookies.get("admin_auth") != "yes":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    remove_hotel_from_owner(owner_id, slug)
    return {"ok": True}

# Admin impersonate hotel manager (no password needed)
@app.get("/api/admin/hotel/{slug}/impersonate")
def api_admin_impersonate(slug: str, request: Request, response: Response):
    if request.cookies.get("admin_auth") != "yes":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)
    response.set_cookie(f"auth_{slug}", "yes", max_age=3600,
                        httponly=True, secure=SECURE_COOKIES, samesite="strict")
    return RedirectResponse(f"/hotel/{slug}/dashboard")


# ===== STAFF AUTH HELPERS =====

def get_auth_role(request: Request, slug: str) -> tuple[bool, str]:
    """
    Returns (is_authorized, role).
    role is one of: "manager", "receptionist", "housekeeping", ""
    Manager cookie (auth_{slug}=yes) → role "manager"
    Staff cookie (staff_{slug}={id}) → look up in DB
    """
    if request.cookies.get(f"auth_{slug}") == "yes":
        return True, "manager"
    staff_id_str = request.cookies.get(f"staff_{slug}", "")
    if staff_id_str:
        try:
            staff = get_staff_by_id(slug, int(staff_id_str))
            if staff and staff["is_active"]:
                return True, staff["role"]
        except (ValueError, TypeError):
            pass
    return False, ""


# ===== STAFF PYDANTIC MODELS =====

class StaffLoginRequest(BaseModel):
    username: str
    password: str

class StaffCreateRequest(BaseModel):
    name: str
    username: str
    password: str
    role: str = "receptionist"

class StaffPasswordReset(BaseModel):
    password: str


# ===== STAFF ENDPOINTS =====

@app.post("/api/hotel/{slug}/staff/login")
@limiter.limit("5/minute")
def staff_login(slug: str, data: StaffLoginRequest, request: Request, response: Response):
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Hotel not found"}, status_code=404)
    staff = get_staff_by_credentials(slug, data.username, data.password)
    if not staff:
        return JSONResponse({"error": "Geçersiz kullanıcı adı veya şifre"}, status_code=401)
    response.set_cookie(
        key=f"staff_{slug}",
        value=str(staff["id"]),
        httponly=True,
        secure=SECURE_COOKIES,
        samesite="lax",
        max_age=86400 * 7  # 7 days
    )
    # Also set hotel_slug cookie so requests endpoints work for staff
    response.set_cookie(
        key="hotel_slug",
        value=slug,
        httponly=True,
        secure=SECURE_COOKIES,
        samesite="lax",
        max_age=86400 * 7
    )
    return {"ok": True, "name": staff["name"], "role": staff["role"]}


@app.get("/api/hotel/{slug}/staff/logout")
def staff_logout(slug: str, response: Response):
    response.delete_cookie(f"staff_{slug}")
    return RedirectResponse(f"/hotel/{slug}/staff/login")


@app.get("/api/hotel/{slug}/me")
def hotel_me(slug: str, request: Request):
    """Returns current user's role for the dashboard to adapt tab visibility."""
    ok, role = get_auth_role(request, slug)
    if not ok:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    if role == "manager":
        name = get_hotel(slug)["name"] + " Yöneticisi"
    else:
        staff_id_str = request.cookies.get(f"staff_{slug}", "")
        try:
            staff = get_staff_by_id(slug, int(staff_id_str))
            name = staff["name"] if staff else "Personel"
        except Exception:
            name = "Personel"
    return {"ok": True, "role": role, "name": name}


@app.get("/api/hotel/{slug}/staff")
def list_staff(slug: str, request: Request):
    ok, role = get_auth_role(request, slug)
    if not ok or role != "manager":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    return get_staff_list(slug)


@app.post("/api/hotel/{slug}/staff")
def add_staff(slug: str, data: StaffCreateRequest, request: Request):
    ok, role = get_auth_role(request, slug)
    if not ok or role != "manager":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    if len(data.password) < 6:
        return JSONResponse({"error": "Şifre en az 6 karakter olmalı"}, status_code=400)
    if data.role not in ("manager", "receptionist", "housekeeping"):
        return JSONResponse({"error": "Geçersiz rol"}, status_code=400)
    staff_id = create_staff(slug, data.name, data.username, data.password, data.role)
    if staff_id is None:
        return JSONResponse({"error": "Bu kullanıcı adı zaten alınmış"}, status_code=409)
    return {"ok": True, "id": staff_id}


@app.delete("/api/hotel/{slug}/staff/{sid}")
def remove_staff(slug: str, sid: int, request: Request):
    ok, role = get_auth_role(request, slug)
    if not ok or role != "manager":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    changed = delete_staff(sid, slug)
    if not changed:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return {"ok": True}


@app.patch("/api/hotel/{slug}/staff/{sid}/password")
def reset_staff_password(slug: str, sid: int, data: StaffPasswordReset, request: Request):
    ok, role = get_auth_role(request, slug)
    if not ok or role != "manager":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    if len(data.password) < 6:
        return JSONResponse({"error": "Şifre en az 6 karakter olmalı"}, status_code=400)
    changed = update_staff_password(sid, slug, data.password)
    if not changed:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return {"ok": True}


# ===== STAFF LOGIN PAGE =====

# ===== STRIPE BILLING =====

@app.post("/api/stripe/webhook")
async def stripe_webhook(request: Request):
    """Receive Stripe events and update hotel plan/subscription status."""
    payload = await request.body()
    sig_header = request.headers.get("stripe-signature", "")

    try:
        event = stripe.Webhook.construct_event(payload, sig_header, STRIPE_WEBHOOK_SECRET)
    except stripe.error.SignatureVerificationError:
        return JSONResponse({"error": "Invalid signature"}, status_code=400)
    except Exception as e:
        print(f"[stripe webhook error] {e}")
        return JSONResponse({"error": "Bir hata oluştu, lütfen tekrar deneyin"}, status_code=400)

    etype = event["type"]
    data  = event["data"]["object"]

    if etype == "checkout.session.completed":
        # New subscription created through Checkout
        customer_id = data.get("customer")
        sub_id      = data.get("subscription")
        slug        = (data.get("metadata") or {}).get("hotel_slug")
        if slug and customer_id:
            # Retrieve subscription to get price info
            plan = "starter"
            try:
                sub = stripe.Subscription.retrieve(sub_id)
                price_id = sub["items"]["data"][0]["price"]["id"]
                plan = STRIPE_PRICE_TO_PLAN.get(price_id, "starter")
            except Exception:
                pass
            update_hotel_stripe(slug,
                stripe_customer_id=customer_id,
                stripe_subscription_id=sub_id or "",
                stripe_status="active",
                plan=plan)

    elif etype == "customer.subscription.updated":
        customer_id = data.get("customer")
        sub_id      = data.get("id")
        status      = data.get("status", "")
        slug = get_hotel_by_stripe_customer(customer_id)
        if slug:
            plan = None
            try:
                price_id = data["items"]["data"][0]["price"]["id"]
                plan = STRIPE_PRICE_TO_PLAN.get(price_id)
            except Exception:
                pass
            update_hotel_stripe(slug,
                stripe_subscription_id=sub_id,
                stripe_status=status,
                plan=plan if status == "active" else None)

    elif etype == "customer.subscription.deleted":
        customer_id = data.get("customer")
        slug = get_hotel_by_stripe_customer(customer_id)
        if slug:
            update_hotel_stripe(slug,
                stripe_subscription_id="",
                stripe_status="canceled",
                plan="trial")

    elif etype == "invoice.payment_failed":
        customer_id = data.get("customer")
        slug = get_hotel_by_stripe_customer(customer_id)
        if slug:
            update_hotel_stripe(slug, stripe_status="past_due")

    return {"ok": True}


@app.post("/api/hotel/{slug}/billing/checkout")
async def billing_checkout(slug: str, request: Request):
    """Create a Stripe Checkout Session for upgrading the plan."""
    ok, role = get_auth_role(request, slug)
    if not ok or role != "manager":
        return JSONResponse({"error": "Unauthorized"}, status_code=403)
    if not stripe.api_key:
        return JSONResponse({"error": "Stripe not configured"}, status_code=503)

    body = await request.json()
    price_id = body.get("price_id", "")
    if not price_id or price_id not in STRIPE_PRICE_TO_PLAN:
        return JSONResponse({"error": "Invalid plan"}, status_code=400)

    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Hotel not found"}, status_code=404)

    base_url = str(request.base_url).rstrip("/")
    try:
        # Reuse existing customer or create a new one
        customer_id = hotel.get("stripe_customer_id") or None
        session = stripe.checkout.Session.create(
            customer=customer_id or stripe.util.UNSET,
            customer_email=None if customer_id else None,
            mode="subscription",
            line_items=[{"price": price_id, "quantity": 1}],
            metadata={"hotel_slug": slug},
            success_url=f"{base_url}/hotel/{slug}/dashboard?billing=success",
            cancel_url=f"{base_url}/hotel/{slug}/dashboard?billing=cancel",
        )
        return {"checkout_url": session.url}
    except stripe.error.StripeError as e:
        print(f"[stripe checkout error] {e}")
        return JSONResponse({"error": "Bir hata oluştu, lütfen tekrar deneyin"}, status_code=500)


@app.get("/api/hotel/{slug}/billing/portal")
async def billing_portal(slug: str, request: Request):
    """Redirect manager to Stripe Customer Portal to manage/cancel subscription."""
    ok, role = get_auth_role(request, slug)
    if not ok or role != "manager":
        return JSONResponse({"error": "Unauthorized"}, status_code=403)
    if not stripe.api_key:
        return JSONResponse({"error": "Stripe not configured"}, status_code=503)

    hotel = get_hotel(slug)
    customer_id = hotel.get("stripe_customer_id") if hotel else ""
    if not customer_id:
        return JSONResponse({"error": "No billing account found"}, status_code=404)

    base_url = str(request.base_url).rstrip("/")
    try:
        session = stripe.billing_portal.Session.create(
            customer=customer_id,
            return_url=f"{base_url}/hotel/{slug}/dashboard",
        )
        return RedirectResponse(session.url, status_code=303)
    except stripe.error.StripeError as e:
        print(f"[stripe portal error] {e}")
        return JSONResponse({"error": "Bir hata oluştu, lütfen tekrar deneyin"}, status_code=500)


@app.get("/api/hotel/{slug}/billing/info")
async def billing_info(slug: str, request: Request):
    """Return current billing state for dashboard UI."""
    ok, role = get_auth_role(request, slug)
    if not ok or role != "manager":
        return JSONResponse({"error": "Unauthorized"}, status_code=403)
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)
    return {
        "plan":            hotel["plan"],
        "stripe_status":   hotel["stripe_status"],
        "has_subscription": bool(hotel.get("stripe_subscription_id")),
        "stripe_configured": bool(stripe.api_key),
        "prices": {
            "starter":  os.getenv("STRIPE_PRICE_BASIC", ""),
            "pro":      os.getenv("STRIPE_PRICE_PRO", ""),
            "premium":  os.getenv("STRIPE_PRICE_ENTERPRISE", ""),
        }
    }


@app.get("/hotel/{slug}/staff/login", response_class=HTMLResponse)
def staff_login_page(slug: str, request: Request):
    hotel = get_hotel(slug)
    if not hotel:
        return HTMLResponse("❌ Otel bulunamadı", status_code=404)
    hotel_name = hotel["name"]
    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Personel Girişi — {hotel_name}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0a0a0a;color:#f0f0f0;font-family:'Segoe UI',sans-serif;
       display:flex;align-items:center;justify-content:center;min-height:100vh}}
  .card{{background:#111;border:1px solid #222;border-radius:16px;padding:40px;width:100%;max-width:380px}}
  h2{{color:#C9A84C;margin-bottom:6px;font-size:22px}}
  .sub{{color:#666;font-size:13px;margin-bottom:28px}}
  label{{font-size:12px;color:#888;text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:5px}}
  input{{width:100%;padding:11px 14px;background:#1a1a1a;border:1px solid #333;border-radius:8px;
         color:#f0f0f0;font-size:15px;outline:none;margin-bottom:16px}}
  input:focus{{border-color:#C9A84C}}
  .btn{{width:100%;padding:13px;background:#C9A84C;color:#000;font-size:15px;font-weight:700;
        border:none;border-radius:8px;cursor:pointer;margin-top:4px}}
  .btn:hover{{background:#d4b85a}}
  .err{{color:#e05555;font-size:13px;margin-top:12px;display:none}}
  .mgr-link{{text-align:center;margin-top:16px;font-size:13px;color:#555}}
  .mgr-link a{{color:#C9A84C;text-decoration:none}}
</style>
</head>
<body>
<div class="card">
  <h2>👤 Personel Girişi</h2>
  <p class="sub">{hotel_name}</p>
  <label>KULLANICI ADI</label>
  <input type="text" id="username" placeholder="kullanici.adi" autocomplete="username">
  <label>ŞİFRE</label>
  <input type="password" id="password" placeholder="••••••••" autocomplete="current-password">
  <button class="btn" onclick="doLogin()">🔑 Giriş Yap</button>
  <div class="err" id="err"></div>
  <div class="mgr-link"><a href="/hotel/{slug}/login">Yönetici olarak giriş yap →</a></div>
</div>
<script>
  const slug = '{slug}';
  document.addEventListener('keydown', e => {{ if(e.key==='Enter') doLogin(); }});
  async function doLogin() {{
    const username = document.getElementById('username').value.trim();
    const password = document.getElementById('password').value;
    if (!username || !password) return;
    const r = await fetch('/api/hotel/' + slug + '/staff/login', {{
      method: 'POST',
      headers: {{'Content-Type':'application/json'}},
      credentials: 'include',
      body: JSON.stringify({{username, password}})
    }});
    const d = await r.json();
    if (d.ok) {{
      window.location.href = '/hotel/' + slug + '/dashboard';
    }} else {{
      const err = document.getElementById('err');
      err.textContent = '❌ ' + (d.error || 'Giriş başarısız');
      err.style.display = 'block';
    }}
  }}
</script>
</body>
</html>""")


# ===== SMART BUFFET =====

@app.get("/hotel/{slug}/buffet", response_class=HTMLResponse)
def buffet_page(slug: str, request: Request):
    """Страница AI-анализа шведского стола. Только для авторизованного персонала."""
    hotel = get_hotel(slug)
    if not hotel:
        return HTMLResponse("❌ Otel bulunamadı", status_code=404)
    ok, role = get_auth_role(request, slug)
    if not ok:
        return RedirectResponse(f"/hotel/{slug}/login", status_code=302)
    return get_buffet_html(hotel["name"], slug)


@app.post("/hotel/{slug}/buffet/analyze")
async def buffet_analyze(slug: str, request: Request, photo: UploadFile = File(...)):
    """
    Принимает фото буфета (multipart form-data, поле 'photo'),
    анализирует через Claude Vision и сохраняет результат.
    Возвращает JSON: {ok, scan_id, result}.
    """
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)
    ok, role = get_auth_role(request, slug)
    if not ok:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    import base64
    raw = await photo.read()
    if len(raw) > 20 * 1024 * 1024:   # 20 MB limit
        return JSONResponse({"error": "File too large (max 20 MB)"}, status_code=413)

    b64 = base64.b64encode(raw).decode()
    media_type = photo.content_type or "image/jpeg"

    try:
        result = await asyncio.to_thread(analyze_buffet_photo, b64, media_type)
    except Exception as e:
        print(f"[buffet analysis error] {e}")
        return JSONResponse({"error": "Bir hata oluştu, lütfen tekrar deneyin"}, status_code=500)

    scan_id = save_buffet_scan(slug, result)
    return {"ok": True, "scan_id": scan_id, "result": result}


@app.get("/api/hotel/{slug}/buffet/latest")
def buffet_latest_api(slug: str, request: Request):
    """Последний скан буфета для дашборда."""
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)
    ok, role = get_auth_role(request, slug)
    if not ok:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    scan = get_buffet_latest(slug)
    return {"scan": scan}


@app.get("/api/hotel/{slug}/buffet/history")
def buffet_history_api(slug: str, request: Request, days: int = 7):
    """История сканов за последние N дней (по умолчанию 7)."""
    hotel = get_hotel(slug)
    if not hotel:
        return JSONResponse({"error": "Not found"}, status_code=404)
    ok, role = get_auth_role(request, slug)
    if not ok:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    days = max(1, min(days, 90))   # clamp 1..90
    scans = get_buffet_history(slug, days)
    return {"scans": scans}
